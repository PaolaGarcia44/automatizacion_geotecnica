"""
Script para crear plantillas Excel de ejemplo

Ejecutar con: python generate_templates.py

Este script crea 3 plantillas Excel básicas para cada categoría
con la estructura necesaria para que el backend funcione.
"""

from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Crear directorio si no existe
TEMPLATES_DIR = Path(__file__).parent / "templates" / "excel"
TEMPLATES_DIR.mkdir(parents=True, exist_ok=True)


def create_header_style():
    """Crea estilo para encabezados"""
    return {
        "font": Font(bold=True, size=12, color="FFFFFF"),
        "fill": PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid"),
        "alignment": Alignment(horizontal="center", vertical="center"),
    }


def create_title_style():
    """Crea estilo para títulos"""
    return {
        "font": Font(bold=True, size=14, color="1F4E78"),
        "alignment": Alignment(horizontal="left", vertical="center"),
    }


def create_border():
    """Crea bordes para celdas"""
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )
    return thin_border


def create_categoria_1():
    """Crea plantilla Categoría 1 (Hasta 3 pisos)"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Proyecto"
    
    # Definir ancho de columnas
    ws.column_dimensions["A"].width = 25
    ws.column_dimensions["B"].width = 20
    ws.column_dimensions["C"].width = 20
    ws.column_dimensions["D"].width = 20
    ws.column_dimensions["E"].width = 20
    ws.column_dimensions["F"].width = 20
    
    # Título principal
    ws["A1"] = "ESTUDIO GEOTÉCNICO - CATEGORÍA 1"
    ws["A1"].font = Font(bold=True, size=14, color="1F4E78")
    
    # Datos básicos
    ws["A3"] = "CATEGORÍA:"
    ws["B3"] = 1
    
    ws["A5"] = "NOMBRE PROYECTO:"
    ws["B5"] = "[Nombre del Proyecto]"
    
    ws["A7"] = "MUNICIPIO:"
    ws["C7"] = "[Municipio]"
    
    ws["A10"] = "FECHA REGISTRO:"
    ws["B10"] = "[Fecha]"
    
    ws["A12"] = "CAMPO N:"
    ws["C12"] = "[Campo N]"
    
    ws["A15"] = "DESCRIPCIÓN:"
    ws["A16:F25"] = "[Descripción del proyecto]"
    
    # Tabla de perforaciones
    ws["A27"] = "PERFORACIONES"
    ws["A27"].font = Font(bold=True, size=11)
    
    # Encabezado de tabla
    headers = ["Número", "Profundidad (m)", "Tipo de Suelo", "Observaciones"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=28, column=col)
        cell.value = header
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    
    # Filas para perforaciones (hasta 10)
    for row in range(29, 39):
        for col in range(1, 5):
            ws.cell(row=row, column=col).border = create_border()
    
    # Guardar
    output_path = TEMPLATES_DIR / "plantilla_categoria_1.xlsx"
    wb.save(output_path)
    print(f"✓ Creada: {output_path}")


def create_categoria_2():
    """Crea plantilla Categoría 2 (Hasta 10 pisos)"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Proyecto"
    
    # Definir ancho de columnas
    for col in ["A", "B", "C", "D", "E", "F"]:
        ws.column_dimensions[col].width = 20
    
    # Título
    ws["A1"] = "ESTUDIO GEOTÉCNICO - CATEGORÍA 2"
    ws["A1"].font = Font(bold=True, size=14, color="1F4E78")
    
    # Datos básicos
    ws["A3"] = "CATEGORÍA:"
    ws["B3"] = 2
    
    ws["A4"] = "NOMBRE PROYECTO:"
    ws["B4"] = "[Nombre del Proyecto]"
    
    ws["A5"] = "CAMPO N:"
    ws["B5"] = "[Campo N]"
    
    ws["A6"] = "MUNICIPIO:"
    ws["C6"] = "[Municipio]"
    
    ws["A7"] = "FECHA REGISTRO:"
    ws["B7"] = "[Fecha]"
    
    ws["A10"] = "DESCRIPCIÓN:"
    ws["A11:E15"] = "[Descripción del proyecto]"
    
    # Tabla de perforaciones
    ws["A18"] = "PERFORACIONES"
    ws["A18"].font = Font(bold=True, size=11)
    
    headers = ["Número", "Profundidad (m)", "Tipo de Suelo", "Observaciones"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=19, column=col)
        cell.value = header
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    
    for row in range(20, 35):
        for col in range(1, 5):
            ws.cell(row=row, column=col).border = create_border()
    
    output_path = TEMPLATES_DIR / "plantilla_categoria_2.xlsx"
    wb.save(output_path)
    print(f"✓ Creada: {output_path}")


def create_categoria_3():
    """Crea plantilla Categoría 3 (Más de 10 pisos)"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Proyecto"
    
    for col in ["A", "B", "C", "D", "E", "F"]:
        ws.column_dimensions[col].width = 20
    
    # Título
    ws["A1"] = "ESTUDIO GEOTÉCNICO - CATEGORÍA 3"
    ws["A1"].font = Font(bold=True, size=14, color="1F4E78")
    
    # Datos básicos
    ws["C2"] = "CATEGORÍA:"
    ws["D2"] = 3
    
    ws["A3"] = "NOMBRE PROYECTO:"
    ws["B3"] = "[Nombre del Proyecto]"
    
    ws["A4"] = "CAMPO N:"
    ws["B4"] = "[Campo N]"
    
    ws["A5"] = "MUNICIPIO:"
    ws["C5"] = "[Municipio]"
    
    ws["A7"] = "TOTAL PERFORACIONES:"
    ws["B7"] = 0
    
    ws["A10"] = "DESCRIPCIÓN:"
    ws["A11:F15"] = "[Descripción del proyecto]"
    
    # Tabla de perforaciones
    ws["A20"] = "PERFORACIONES"
    ws["A20"].font = Font(bold=True, size=11)
    
    headers = ["Número", "Profundidad (m)", "Tipo de Suelo", "Observaciones"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=21, column=col)
        cell.value = header
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    
    for row in range(22, 42):
        for col in range(1, 5):
            ws.cell(row=row, column=col).border = create_border()
    
    output_path = TEMPLATES_DIR / "plantilla_categoria_3.xlsx"
    wb.save(output_path)
    print(f"✓ Creada: {output_path}")


def main():
    """Función principal"""
    print("\n" + "=" * 60)
    print("Generando plantillas Excel...")
    print("=" * 60)
    
    try:
        create_categoria_1()
        create_categoria_2()
        create_categoria_3()
        
        print("\n" + "=" * 60)
        print("✓ Todas las plantillas han sido creadas exitosamente")
        print(f"Ubicación: {TEMPLATES_DIR}")
        print("=" * 60 + "\n")
        
    except Exception as e:
        print(f"\n✗ Error al crear plantillas: {str(e)}\n")


if __name__ == "__main__":
    main()
