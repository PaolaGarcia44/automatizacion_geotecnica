"""Document orchestration service for the Excel templates."""

import logging
import math
import random
import re
import shutil
import unicodedata
from concurrent.futures import ThreadPoolExecutor, as_completed
from copy import copy
from datetime import date, datetime, timedelta
from pathlib import Path
from typing import Dict, List, Optional
from urllib.parse import quote
from uuid import uuid4
from zipfile import ZIP_DEFLATED, ZIP_STORED, ZipFile

from openpyxl import load_workbook
from openpyxl.cell.cell import MergedCell
from openpyxl.formula.translate import Translator
from openpyxl.styles import Alignment, PatternFill

from app.core.config import settings
from app.core.constants import get_pisos_config
from app.utils.common import parse_fecha as _parse_fecha
from app.services.excel_service import excel_service
from app.services.word_service import word_service
from app.services.profile_3d_service import profile_3d_service

logger = logging.getLogger(__name__)


class DocumentService:
    """Central service for generating geotechnical Excel files."""
    
    def __init__(self):
        self.excel_service = excel_service

    def _slugify_filename(self, value: str, fallback: str) -> str:
        text = unicodedata.normalize("NFD", value or "").encode("ascii", "ignore").decode("ascii")
        text = re.sub(r"[^A-Za-z0-9]+", "_", text).strip("_")
        return text or fallback

    def _sanitize_archive_name(self, name: str, max_length: int = 100) -> str:
        if not name:
            return name
        # preserve extension
        try:
            parts = name.rsplit('.', 1)
            base = parts[0]
            ext = '.' + parts[1] if len(parts) == 2 else ''
        except Exception:
            base, ext = name, ''

        # normalize and remove non-ascii
        base_norm = unicodedata.normalize('NFD', base).encode('ascii', 'ignore').decode('ascii')
        # replace problematic chars with underscores
        base_norm = re.sub(r'[^A-Za-z0-9\-_ ]+', '', base_norm).strip()
        base_norm = re.sub(r'\s+', ' ', base_norm)
        # shorten if necessary
        if len(base_norm) > max_length:
            base_norm = base_norm[:max_length].rstrip()
        # replace spaces with single spaces (keep readable) and finally ensure no leading/trailing
        safe_name = base_norm.strip()
        # restore a compact filename
        safe_name = safe_name[:max_length]
        return f"{safe_name}{ext}"

    def _prepare_profile_file(
        self,
        source_template: Path,
        project_id: str,
        project_value,
        fecha_value,
        fecha_original=None,
        perforaciones: Optional[List[Dict]] = None,
    ) -> Path:
        output_name = f"{project_id}_{source_template.stem}.xlsx"
        output_path = self.excel_service.generated_dir / output_name

        shutil.copy2(source_template, output_path)

        workbook = load_workbook(output_path)
        worksheet = workbook.active
        worksheet["D1"] = project_value

        def _truncate_depth(value) -> Optional[int]:
            if value in (None, ""):
                return None
            text = str(value).strip().replace(",", ".")
            try:
                return max(1, int(float(text)))
            except Exception:
                return None

        def _profile_block_ranges(ws, start_row: int = 5):
            ranges = []
            handled_rows = set()
            for row_number in range(start_row, ws.max_row + 1):
                if row_number in handled_rows:
                    continue

                d_cell = ws[f"D{row_number}"]
                if isinstance(d_cell, MergedCell):
                    continue

                containing_merges = [
                    merged_range
                    for merged_range in ws.merged_cells.ranges
                    if merged_range.min_row <= row_number <= merged_range.max_row
                    and merged_range.min_col <= 4 <= merged_range.max_col
                ]

                if not containing_merges and d_cell.value in (None, ""):
                    continue

                row_end = max((merged_range.max_row for merged_range in containing_merges), default=row_number)
                ranges.append((row_number, row_end))
                handled_rows.update(range(row_number, row_end + 1))

            return ranges

        def _clear_cell_if_writable(ws, cell_ref: str):
            cell = ws[cell_ref]
            if isinstance(cell, MergedCell):
                return
            cell.value = None
            cell.fill = PatternFill(fill_type=None)

        def _apply_profile_j_formulas(ws):
            for start_row, end_row in _profile_block_ranges(ws):
                span_rows = end_row - start_row + 1
                if span_rows == 1:
                    formula = f"=I{start_row}"
                else:
                    i_terms = "+".join([f"I{row_number}" for row_number in range(start_row, end_row + 1)])
                    formula = f"=({i_terms})/{span_rows}"
                ws[f"J{start_row}"] = formula
                ws[f"D{start_row}"].number_format = "0.00"


        profile_name = source_template.name.upper()
        if any(tag in profile_name for tag in ("6M", "15M", "25M")):
            max_depth_row = 0
            for row_number in range(5, worksheet.max_row + 1):
                c_value = worksheet[f"C{row_number}"].value
                if c_value in (None, ""):
                    continue
                try:
                    float(str(c_value).replace(",", "."))
                    max_depth_row = row_number
                except Exception:
                    continue

            layers = list(perforaciones or [])
            mirror_columns = ["J", "K", "L", "M"]
            mirror_sources = [5, 6]
            mirror_template_values = {
                source_row: {
                    column: worksheet[f"{column}{source_row}"].value for column in mirror_columns
                }
                for source_row in mirror_sources
            }

            def _copy_formula_or_value(source_cell_ref: str, target_cell_ref: str):
                source_value = mirror_template_values[int("".join(ch for ch in source_cell_ref if ch.isdigit()))][
                    "".join(ch for ch in source_cell_ref if ch.isalpha())
                ]
                if isinstance(source_value, str) and source_value.startswith("="):
                    try:
                        return Translator(source_value, origin=source_cell_ref).translate_formula(target_cell_ref)
                    except Exception:
                        return source_value
                return source_value

            def _depth_floor(value) -> Optional[int]:
                if value in (None, ""):
                    return None
                text = str(value).strip().replace(",", ".")
                try:
                    d = float(text)
                    return math.ceil(d) if d > 0 else None
                except Exception:
                    return None

            depth_rows = list(range(5, max_depth_row + 1))
            if layers:
                cleaned_layers = []
                previous_floor = 0
                for index, layer in enumerate(layers):
                    depth_floor = _depth_floor(layer.get("profundidad_z"))
                    if depth_floor is None:
                        depth_floor = previous_floor + 1
                    if depth_floor <= previous_floor:
                        depth_floor = previous_floor + 1

                    cleaned_layers.append((layer, depth_floor))
                    previous_floor = depth_floor

                total_visible_rows = len(depth_rows)
                # profundidad_z = depth where the layer STARTS.
                # Span = next layer's ceil(depth) - this layer's ceil(depth).
                # Last layer fills remaining rows.
                spans: List[int] = []
                prev_depth_floor = 0
                for index, (_, depth_floor) in enumerate(cleaned_layers):
                    if index + 1 < len(cleaned_layers):
                        span_rows = max(1, depth_floor - prev_depth_floor)
                    else:
                        span_rows = max(1, total_visible_rows - sum(spans))
                    spans.append(span_rows)
                    prev_depth_floor = depth_floor

                if spans and sum(spans) > total_visible_rows:
                    overflow = sum(spans) - total_visible_rows
                    spans[-1] = max(1, spans[-1] - overflow)

                # Rebuild only the layer area so the number of visible layers matches the input.
                for merged_range in list(worksheet.merged_cells.ranges):
                    if merged_range.min_row > max_depth_row or merged_range.max_row < 5:
                        continue
                    if (
                        merged_range.min_col in {2, 4, 10, 11, 12, 13}
                        or merged_range.max_col in {2, 4, 10, 11, 12, 13}
                    ):
                        worksheet.unmerge_cells(str(merged_range))

                for row_number in depth_rows:
                    worksheet[f"B{row_number}"] = None
                    worksheet[f"D{row_number}"] = None
                    worksheet[f"B{row_number}"].fill = PatternFill(fill_type=None)
                    worksheet[f"D{row_number}"].fill = PatternFill(fill_type=None)

                for row_number in depth_rows:
                    for column in mirror_columns:
                        worksheet[f"{column}{row_number}"].fill = PatternFill(fill_type=None)

                # Clear K, L, M, N in entire depth range — only anchor rows will get values
                for row_number in depth_rows:
                    _clear_cell_if_writable(worksheet, f"K{row_number}")
                    _clear_cell_if_writable(worksheet, f"L{row_number}")
                    _clear_cell_if_writable(worksheet, f"M{row_number}")
                    _clear_cell_if_writable(worksheet, f"N{row_number}")

                current_row = depth_rows[0]
                for index, ((layer, depth_floor), span_rows) in enumerate(zip(cleaned_layers, spans)):
                    row_end = min(max_depth_row, current_row + span_rows - 1)
                    if index == len(spans) - 1:
                        row_end = max_depth_row

                    soil_name = self.excel_service._clean_soil_text(
                        layer.get("descripcion_suelo"),
                        layer.get("color_predominante"),
                    )
                    if not soil_name:
                        soil_name = self.excel_service._clean_soil_text(
                            layer.get("tipo_suelo_principal"),
                            layer.get("color_predominante"),
                        )

                    fill, font_color = self.excel_service._soil_style_from_color(layer.get("color_predominante"))

                    b_anchor = worksheet[f"B{current_row}"]
                    d_anchor = worksheet[f"D{current_row}"]

                    b_anchor.value = soil_name
                    b_anchor.fill = fill
                    b_anchor.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                    b_font = copy(b_anchor.font)
                    b_font.color = font_color
                    b_anchor.font = b_font

                    d_anchor.value = float(span_rows)
                    d_anchor.number_format = "0.00"
                    d_anchor.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

                    source_row = mirror_sources[0] if index == 0 else mirror_sources[-1]
                    for column in mirror_columns:
                        target_anchor = worksheet[f"{column}{current_row}"]
                        if column == 'J':
                            i_terms = "+".join([f"I{row_number}" for row_number in range(current_row, row_end + 1)])
                            target_anchor.value = f"=({i_terms})/{span_rows}"
                        elif column == 'K':
                            target_anchor.value = 0
                        elif column == 'L':
                            target_anchor.value = 0
                        elif column == 'M':
                            if "15M" in profile_name or "25M" in profile_name:
                                target_anchor.value = 0  # Cohesión for 15M/25M
                            else:
                                # 6M: M is di/Ni = espesor / N_promedio_estrato
                                target_anchor.value = f"=D{current_row}/J{current_row}"
                                target_anchor.number_format = "0.000"
                        else:
                            target_anchor.value = _copy_formula_or_value(f"{column}{source_row}", f"{column}{current_row}")
                        target_anchor.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

                    # 15M/25M: di/Ni is in column N (=D/K). 6M: di/Ni is in column M (=D/J), N is unused.
                    if "15M" in profile_name or "25M" in profile_name:
                        n_anchor = worksheet[f"N{current_row}"]
                        if not isinstance(n_anchor, MergedCell):
                            n_anchor.value = f"=IF(K{current_row}=0,0,D{current_row}/K{current_row})"
                            n_anchor.number_format = "0.000"
                            n_anchor.alignment = Alignment(horizontal="center", vertical="center")

                    if row_end > current_row:
                        worksheet.merge_cells(start_row=current_row, start_column=2, end_row=row_end, end_column=2)
                        worksheet.merge_cells(start_row=current_row, start_column=4, end_row=row_end, end_column=4)
                        for column_letter in mirror_columns:
                            column_index = ord(column_letter) - ord("A") + 1
                            worksheet.merge_cells(start_row=current_row, start_column=column_index, end_row=row_end, end_column=column_index)

                    current_row = row_end + 1
                    if current_row > max_depth_row:
                        break

                for row_number in range(current_row, max_depth_row + 1):
                    _clear_cell_if_writable(worksheet, f"B{row_number}")
                    _clear_cell_if_writable(worksheet, f"D{row_number}")

                # If there are more capas than visible depth rows, collapse the extras into the last layer.
                if "15M" in profile_name or "25M" in profile_name:
                    worksheet["M1"] = "FECHA"
                    worksheet["N1"] = fecha_value
                else:
                    worksheet["M1"] = fecha_value
                    worksheet["N1"] = None
            else:
                for row_number in depth_rows:
                    _clear_cell_if_writable(worksheet, f"B{row_number}")
                    _clear_cell_if_writable(worksheet, f"D{row_number}")
                    _clear_cell_if_writable(worksheet, f"K{row_number}")
                    _clear_cell_if_writable(worksheet, f"L{row_number}")
                if "15M" in profile_name or "25M" in profile_name:
                    worksheet["M1"] = "FECHA"
                    worksheet["N1"] = fecha_value
                else:
                    worksheet["M1"] = fecha_value
                    worksheet["N1"] = None
        else:
            worksheet["N1"] = fecha_value

        if any(tag in profile_name for tag in ("6M", "15M", "25M")):
            _apply_profile_j_formulas(worksheet)
        workbook.save(output_path)
        workbook.close()

        return output_path
    
    def generate_documents(
        self,
        proyecto_ubicacion: str,
        cliente: Optional[str],
        fecha_registro,
        sondeo: Optional[str],
        pisos: int,
        perforaciones: Optional[List[Dict]] = None,
        parametros: Optional[List[Dict]] = None,
        template_id: Optional[str] = None,
        template_ids: Optional[List[str]] = None,
        images_dir: Optional[Path] = None,
        project_id: Optional[str] = None,
        clasificacion_suelo: Optional[str] = None,
        clasificaciones_por_lab: Optional[Dict[str, str]] = None,
        municipio_word: Optional[str] = None,
        word_template_filename: Optional[str] = None,
        valores_laboratorio: Optional[Dict] = None,
        valores_laboratorio_por_lab: Optional[Dict[str, Dict]] = None,
        capacidad_portante: Optional[Dict] = None,
    ) -> Dict:
        """Generate the Excel file from the selected template."""        
        try:
            if not project_id:
                project_id = str(uuid4())[:8]
            timestamp = datetime.now().isoformat()

            perforaciones_to_use = perforaciones or []

            # Decide plantilla and defaults either from an explicit template_id or from pisos.
            nPisos = int(pisos or 0)
            requested_template = str(template_id).strip() if template_id is not None and str(template_id).strip() else None

            def _default_perforaciones_for_template(template_key: str) -> List[Dict]:
                if template_key == '1':
                    return [
                        {"profundidad_z": 6, "gamma": 15, "n_campo_spt": 0, "cohesion_c": None, "descripcion_suelo": ""},
                        {"profundidad_z": 6, "gamma": 16, "n_campo_spt": 0, "cohesion_c": None, "descripcion_suelo": ""},
                        {"profundidad_z": 6, "gamma": 17, "n_campo_spt": 0, "cohesion_c": None, "descripcion_suelo": ""},
                    ]
                if template_key in {'2', '3'}:
                    return [
                        {"profundidad_z": 15, "gamma": None, "n_campo_spt": 0, "cohesion_c": None, "descripcion_suelo": ""},
                        {"profundidad_z": 15, "gamma": None, "n_campo_spt": 0, "cohesion_c": None, "descripcion_suelo": ""},
                        {"profundidad_z": 15, "gamma": None, "n_campo_spt": 0, "cohesion_c": None, "descripcion_suelo": ""},
                        {"profundidad_z": 15, "gamma": None, "n_campo_spt": 0, "cohesion_c": None, "descripcion_suelo": ""},
                    ]
                return []

            if requested_template in {'1', '2', '3'}:
                selected_template = requested_template
                default_perforaciones = perforaciones_to_use or _default_perforaciones_for_template(selected_template)
            elif requested_template == '8':
                selected_template = requested_template
                default_perforaciones = perforaciones_to_use
            elif requested_template in {'4', '5', '6'}:
                selected_template = requested_template
                default_perforaciones = perforaciones_to_use
            elif not perforaciones_to_use:
                selected_template = '8'
                default_perforaciones = perforaciones_to_use
            else:
                selected_template = '8'
                default_perforaciones = perforaciones_to_use

            # If the user provided soil menu values, keep the visible text focused on
            # the soil type; the predominant color will be rendered as cell fill.
            for row_data in default_perforaciones:
                tipo_suelo = row_data.get('tipo_suelo_principal')
                descripcion_suelo = row_data.get('descripcion_suelo')

                if not descripcion_suelo:
                    row_data['descripcion_suelo'] = str(tipo_suelo or '').strip()

            # Fill gamma values sequentially when they are missing.
            # The first soil layer starts at 15 and each additional layer increments by 1.
            # This applies to all templates so the gamma column stays aligned with the
            # number of layers the user entered from the soil menus.
            for index, row_data in enumerate(default_perforaciones):
                if row_data.get('gamma') in (None, ''):
                    row_data['gamma'] = 15 + index

            # Only write Proyecto and Fecha to the template; leave all other cells unchanged
            # Force proyecto_ubicacion to uppercase for the copy
            proyecto_upper = proyecto_ubicacion.upper() if isinstance(proyecto_ubicacion, str) else proyecto_ubicacion

            # Calculate fecha_registro + 20 days for cell C7. Support date or ISO string inputs.
            # Parseo centralizado — fuente única de verdad (common.parse_fecha)
            fecha_obj = _parse_fecha(fecha_registro)
            fecha_c7 = (fecha_obj + timedelta(days=20)) if fecha_obj else None

            excel_data = {
                "proyecto_ubicacion": proyecto_upper,
                "cliente_display": f"Cliente: {cliente}" if cliente else None,
                "fecha_registro": fecha_c7 if fecha_c7 is not None else fecha_registro,
                "fecha_registro_original": fecha_obj if fecha_obj is not None else fecha_registro,
                "cliente": cliente,
                "sondeo": sondeo,
                "pisos": nPisos,
                # static label required by UI: A5 should read 'Parámetro:'
                "parametro_label": "Parámetro:",
            }

            requested_templates = [str(item).strip() for item in (template_ids or []) if str(item).strip()]
            if requested_templates:
                # Batch mode: include the capacity template, the correlation template
                # for the floor range, plus the LABORATORIO set.
                # constants.get_pisos_config es la fuente única de verdad.
                capacity_template_id = '8'
                _pisos_cfg = get_pisos_config(nPisos)
                correlation_template_id = _pisos_cfg["template_id"]
                laboratorio_template_ids = ['4', '5', '6', '7'] if nPisos > 3 else ['4', '5', '6']
                # Template '9' (INCONFINADO) se mueve a Fase 3 para eliminar
                # la condición de carrera con template '4' (lab_state).
                batch_templates = [capacity_template_id, correlation_template_id, *laboratorio_template_ids]
                client_slug = self._slugify_filename(str(cliente or '').strip(), 'cliente')[:30]
                project_slug = self._slugify_filename(proyecto_upper, 'proyecto')[:30]
                municipio_slug = self._slugify_filename(str(municipio_word or '').strip(), '')[:20] if municipio_word else ''
                zip_base_name = f"{client_slug} - {project_slug}{' - ' + municipio_slug if municipio_slug else ''}.zip"
                zip_path = self.excel_service.generated_dir / zip_base_name

                output_files = []
                perfil_added = False
                perfil_error = None
                asentamientos_template_ids = ['10', '11']
                asentamientos_template_names = {
                    '10': 'ASENTAMIENTOS ZAPATAS 300.xlsx',
                    '11': 'ASENTAMIENTOS ZAPATAS 800.xlsx',
                }
                p_template_entries = [
                    ('12', 'P-1.xls', 'P-1'),
                    ('13', 'P-2.xls', 'P-2'),
                    ('14', 'P-3.xls', 'P-3'),
                ]
                if nPisos > 3:
                    p_template_entries.append(('15', 'P-4.xls', 'P-4'))

                # List of all template ids included in the ZIP (for metadata)
                # '9' se incluye en el listado de IDs aunque corra en Fase 3
                batch_templates_with_asentamientos = (
                    list(batch_templates) + ['9'] + list(asentamientos_template_ids)
                    + [str(tpl_id) for tpl_id, _, _ in p_template_entries]
                )

                # Archive names for batch templates — built once, reused across phases
                _archive_name_map = {
                    '1': 'CORRELACIÓN GEOTÉCNICA DE PARÁMETROS GEOMECÁNICOS.xlsx',
                    '2': 'CORRELACIÓN GEOTÉCNICA DE PARÁMETROS GEOMECÁNICOS.xlsx',
                    '3': 'CORRELACIÓN GEOTÉCNICA DE PARÁMETROS GEOMECÁNICOS.xlsx',
                    '4': 'LABORATORIO - FORMULAS 1.xlsx',
                    '5': 'LABORATORIO - FORMULAS 2.xlsx',
                    '6': 'LABORATORIO - FORMULAS 3.xlsx',
                    '7': 'LABORATORIO - FORMULAS 4.xlsx',
                }

                def _batch_archive_name(tpl_id: str, generated_file: Path) -> str:
                    if tpl_id == str(capacity_template_id):
                        return settings.TEMPLATES_CONFIG.get(tpl_id, generated_file.name)
                    return _archive_name_map.get(
                        tpl_id,
                        settings.TEMPLATES_CONFIG.get(tpl_id, generated_file.name),
                    )

                def _cls_vl(tpl_id: str):
                    cls = (clasificaciones_por_lab or {}).get(tpl_id) or clasificacion_suelo
                    vl = (valores_laboratorio_por_lab or {}).get(tpl_id) or valores_laboratorio
                    return cls, vl

                def _gen_batch_excel(tpl_id: str) -> tuple:
                    cls, vl = _cls_vl(tpl_id)
                    f = self.excel_service.generate_excel(
                        template_id=tpl_id,
                        project_id=project_id,
                        data=excel_data,
                        perforaciones=default_perforaciones,
                        parametros=parametros or [],
                        clasificacion_suelo=cls,
                        valores_laboratorio=vl,
                        capacidad_portante=capacidad_portante if tpl_id == '8' else None,
                    )
                    return tpl_id, f

                def _gen_asentamiento(tpl_id: str) -> tuple:
                    f = self.excel_service.generate_excel(
                        template_id=tpl_id,
                        project_id=project_id,
                        data=excel_data,
                        perforaciones=default_perforaciones,
                        parametros=parametros or [],
                        capacidad_portante=capacidad_portante,
                    )
                    return tpl_id, f

                # Scan images dir once so each P template doesn't repeat the glob
                _all_images: list = []
                if images_dir and images_dir.exists():
                    try:
                        _all_images = [
                            f for f in images_dir.iterdir()
                            if f.is_file() and f.suffix.lower() in {'.jpg', '.jpeg', '.png'}
                        ]
                    except Exception:
                        pass

                def _gen_p_template(tpl_id: str, pt_archive: str, pt_sondeo: str) -> tuple:
                    excel_data_pt = dict(excel_data, sondeo=pt_sondeo)
                    _prefix = f"{pt_sondeo} M-".upper()
                    _p_photos = sorted(
                        [f for f in _all_images if f.name.upper().startswith(_prefix)],
                        key=lambda p: p.name,
                    )
                    f = self.excel_service.generate_excel(
                        template_id=tpl_id,
                        project_id=project_id,
                        data=excel_data_pt,
                        perforaciones=default_perforaciones,
                        parametros=parametros or [],
                        photo_paths=_p_photos if _p_photos else None,
                    )
                    return tpl_id, pt_archive, f

                # zip_entries collects (file_path, archive_name, compress_type) in order
                zip_entries: list = []

                # Pre-seed lab state so template '8' (CAPACIDAD PORTANTE) can read
                # n_golpes even though it runs before the LABORATORIO templates.
                # Template '4' will overwrite this with its actual K14 value later.
                _pre_n = random.choice([14, 15, 16])
                self.excel_service._save_lab_state({'n_golpes': _pre_n, 'lp': 22})

                # ── PHASE 1: Template '8' (sequential) ───────────────────────────
                # Must run before '4' overwrites lab_state.
                _t8_id, _t8_file = _gen_batch_excel('8')
                output_files.append(_t8_file)
                zip_entries.append((
                    _t8_file,
                    self._sanitize_archive_name(_batch_archive_name('8', _t8_file)),
                    ZIP_STORED,
                ))

                # ── PHASE 2: Remaining batch templates in parallel ────────────────
                # ('1/2/3', '4', '5', '6', '7') — independent of each other.
                # '4' guarda lab_state; '1/2/3' guardan SPT state.
                # '9' (INCONFINADO) corre en Fase 3 para leer lab_state correcto.
                remaining_batch = [t for t in batch_templates if t != '8']
                phase2_results: dict = {}
                _p2_workers = min(len(remaining_batch), 4)
                with ThreadPoolExecutor(max_workers=_p2_workers) as _ex2:
                    _p2_futures = {_ex2.submit(_gen_batch_excel, t): t for t in remaining_batch}
                    for _fut in as_completed(_p2_futures):
                        _tid, _fpath = _fut.result()
                        phase2_results[_tid] = _fpath

                for t in remaining_batch:
                    f = phase2_results[t]
                    output_files.append(f)
                    zip_entries.append((
                        f,
                        self._sanitize_archive_name(_batch_archive_name(t, f)),
                        ZIP_STORED,
                    ))

                # ── PHASE 3: All remaining documents in parallel ──────────────────
                # lab_state written by '4' and SPT state written by '1/2/3' are now stable.
                # Asentamientos ('10','11'), P templates, PERFIL SUELO, SVG 3D, Word.

                # Nombre del PERFIL DEL SUELO derivado de constants.get_pisos_config
                _perfil_excel_name = _pisos_cfg["perfil_name"]
                _perfil_template = self.excel_service.templates_dir / _perfil_excel_name
                svg_3d_filename = f"{project_id}_PERFIL_3D.svg"
                svg_3d_path = self.excel_service.generated_dir / svg_3d_filename

                def _gen_perfil_excel():
                    return self._prepare_profile_file(
                        source_template=_perfil_template,
                        project_id=project_id,
                        project_value=proyecto_upper,
                        fecha_value=excel_data.get("fecha_registro_original"),
                        fecha_original=excel_data.get("fecha_registro_original"),
                        perforaciones=default_perforaciones,
                    )

                def _gen_svg_3d():
                    profile_3d_service.save_profile_3d(
                        perforaciones=default_perforaciones,
                        output_path=svg_3d_path,
                        project_name=proyecto_upper,
                        sondeo=sondeo or "P-1",
                        fecha=excel_data.get("fecha_registro_original"),
                        pisos=nPisos,
                    )

                def _gen_word():
                    return word_service.generate_informe(
                        project_id=project_id,
                        proyecto_ubicacion=proyecto_upper,
                        fecha_registro=excel_data.get("fecha_registro_original") or fecha_registro,
                        municipio_word=municipio_word or None,
                        template_filename=word_template_filename or None,
                        cliente=cliente,
                        sondeo=sondeo,
                        pisos=nPisos,
                        perforaciones=default_perforaciones,
                        clasificacion_suelo=clasificacion_suelo,
                    )

                def _gen_inconfinado():
                    # Template '9' corre en Fase 3 (después de Fase 2) para que
                    # lab_state escrito por template '4' esté disponible sin
                    # condición de carrera.
                    cls, vl = _cls_vl('9')
                    return '9', self.excel_service.generate_excel(
                        template_id='9',
                        project_id=project_id,
                        data=excel_data,
                        perforaciones=default_perforaciones,
                        parametros=parametros or [],
                        clasificacion_suelo=cls,
                        valores_laboratorio=vl,
                    )

                # Phase 3 results keyed by (type, id)
                _p3_asentamientos: dict = {}
                _p3_p_templates: dict = {}
                _p3_perfil_excel = None
                _p3_word_result = None
                _p3_svg_ok = False
                _p3_inconfinado_file = None

                # P templates run sequentially in a single thread to avoid
                # concurrent Excel.exe processes saving to the same directory
                # (Excel creates ~$ temp files that conflict across instances).
                def _gen_p_templates_sequential() -> dict:
                    results = {}
                    for _pt_id, _pt_arch, _pt_son in p_template_entries:
                        _tid, _tarch, _tfile = _gen_p_template(_pt_id, _pt_arch, _pt_son)
                        results[_tid] = (_tarch, _tfile)
                    return results

                _p3_workers = min(
                    len(asentamientos_template_ids) + 5,  # asentamientos + p_seq + perfil + svg + word + inconfinado
                    8,
                )
                with ThreadPoolExecutor(max_workers=_p3_workers) as _ex3:
                    # Asentamientos
                    _asen_futures = {
                        _ex3.submit(_gen_asentamiento, at): at
                        for at in asentamientos_template_ids
                    }
                    # P templates — one thread, sequential inside to avoid Excel temp-file conflicts
                    _p_seq_fut = _ex3.submit(_gen_p_templates_sequential)
                    # PERFIL DEL SUELO Excel
                    _perfil_fut = _ex3.submit(_gen_perfil_excel) if _perfil_template.exists() else None
                    # SVG 3D
                    _svg_fut = _ex3.submit(_gen_svg_3d)
                    # Word
                    _word_fut = _ex3.submit(_gen_word)
                    # INCONFINADO — Fase 3 para leer lab_state correcto de template '4'
                    _inc_fut = _ex3.submit(_gen_inconfinado)

                    # Collect asentamientos
                    for _fut in as_completed(_asen_futures):
                        _at_id, _at_file = _fut.result()
                        _p3_asentamientos[_at_id] = _at_file

                    # Collect P templates (all results from sequential runner)
                    _p3_p_templates = _p_seq_fut.result()

                    # Collect PERFIL DEL SUELO Excel
                    if _perfil_fut is not None:
                        try:
                            _p3_perfil_excel = _perfil_fut.result()
                        except Exception:
                            logger.warning("No se pudo generar el PERFIL DEL SUELO Excel", exc_info=True)
                    else:
                        logger.warning("Plantilla no encontrada: %s", _perfil_template)

                    # Collect SVG 3D
                    try:
                        _svg_fut.result()
                        _p3_svg_ok = True
                    except Exception:
                        perfil_error = "No se pudo generar el perfil estratigráfico 3D"
                        logger.warning(perfil_error, exc_info=True)

                    # Collect Word
                    try:
                        _p3_word_result = _word_fut.result()
                    except Exception:
                        logger.warning("No se pudo generar el informe Word", exc_info=True)

                    # Collect INCONFINADO
                    try:
                        _, _p3_inconfinado_file = _inc_fut.result()
                    except Exception:
                        logger.warning("No se pudo generar INCONFINADO", exc_info=True)

                # Add phase 3 results to zip_entries in original order

                # INCONFINADO
                if _p3_inconfinado_file is not None:
                    output_files.append(_p3_inconfinado_file)
                    zip_entries.append((
                        _p3_inconfinado_file,
                        self._sanitize_archive_name(
                            settings.TEMPLATES_CONFIG.get('9', _p3_inconfinado_file.name)
                        ),
                        ZIP_STORED,
                    ))

                for at in asentamientos_template_ids:
                    _at_f = _p3_asentamientos[at]
                    output_files.append(_at_f)
                    zip_entries.append((
                        _at_f,
                        self._sanitize_archive_name(asentamientos_template_names[at]),
                        ZIP_STORED,
                    ))

                for pt_id, pt_archive, _ in p_template_entries:
                    _pt_arch, _pt_f = _p3_p_templates[pt_id]
                    output_files.append(_pt_f)
                    zip_entries.append((
                        _pt_f,
                        self._sanitize_archive_name(pt_archive),
                        ZIP_STORED,
                    ))

                if _p3_perfil_excel is not None:
                    output_files.append(_p3_perfil_excel)
                    zip_entries.append((_p3_perfil_excel, _perfil_excel_name, ZIP_STORED))

                if _p3_svg_ok:
                    output_files.append(svg_3d_path)
                    zip_entries.append((svg_3d_path, "PERFIL ESTRATIGRAFICO 3D.svg", ZIP_DEFLATED))
                    # Incluir PNG junto al SVG (para pegar en Word directamente)
                    png_3d_path = svg_3d_path.with_suffix(".png")
                    if png_3d_path.exists():
                        output_files.append(png_3d_path)
                        zip_entries.append((png_3d_path, "PERFIL ESTRATIGRAFICO 3D.png", ZIP_DEFLATED))
                    perfil_added = True

                if _p3_word_result is not None:
                    _inf_file, _inf_archive = _p3_word_result
                    output_files.append(_inf_file)
                    zip_entries.append((
                        _inf_file,
                        self._sanitize_archive_name(_inf_archive),
                        ZIP_STORED,
                    ))

                # ── Write ZIP once with all collected files ───────────────────────
                with ZipFile(zip_path, 'w') as zip_file:
                    for _zf_path, _zf_name, _zf_compress in zip_entries:
                        zip_file.write(_zf_path, arcname=_zf_name, compress_type=_zf_compress)

                    # Include uploaded images in a subfolder inside the ZIP if present
                    try:
                        if images_dir and images_dir.exists():
                            for img in sorted(images_dir.iterdir()):
                                if img.is_file():
                                    zip_file.write(img, arcname=f"imagenes/{img.name}", compress_type=ZIP_STORED)
                                    output_files.append(img)
                    except Exception:
                        logger.debug("No se pudieron agregar imágenes al ZIP", exc_info=True)

                if not perfil_added:
                    raise RuntimeError(perfil_error or "No se pudo generar el perfil estratigráfico")

                return {
                    "success": True,
                    "message": "Documentos generados exitosamente en paquete ZIP",
                    "project_id": project_id,
                    "files": [str(file_path) for file_path in output_files] + [str(zip_path)],
                    "download_url": f"/api/download/{quote(zip_path.name)}",
                    "timestamp": timestamp,
                    "template_id": ",".join(batch_templates_with_asentamientos),
                    "proyecto_ubicacion": proyecto_upper,
                }

            logger.info("Iniciando generación de documentos para proyecto: %s", proyecto_ubicacion)

            excel_file = self.excel_service.generate_excel(
                template_id=selected_template,
                project_id=project_id,
                data=excel_data,
                perforaciones=default_perforaciones,
                parametros=parametros or [],
            )

            logger.info("Documentos generados exitosamente. Proyecto ID: %s", project_id)

            return {
                "success": True,
                "message": "Documentos generados exitosamente",
                "project_id": project_id,
                "files": [str(excel_file)],
                "download_url": f"/api/download/{quote(excel_file.name)}",
                "timestamp": timestamp,
                "template_id": selected_template,
                "proyecto_ubicacion": proyecto_upper,
            }

        except Exception as e:
            logger.exception("Error al generar documentos: %s", str(e))
            return {
                "success": False,
                "message": f"Error al generar documentos: {str(e)}",
                "project_id": None,
            }

    def get_available_templates_status(self) -> Dict:
        """Get the status of the configured templates."""
        return self.excel_service.verify_templates()


# Instancia global del servicio
document_service = DocumentService()
