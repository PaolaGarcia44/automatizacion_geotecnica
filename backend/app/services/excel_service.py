"""Excel service for the two geotechnical templates."""

from __future__ import annotations

import json
import logging
import random
import re
import shutil
import unicodedata
from copy import copy
from datetime import date, datetime
from pathlib import Path
from tempfile import NamedTemporaryFile
from typing import List, Optional
from xml.etree import ElementTree as ET
from zipfile import ZIP_DEFLATED, ZipFile

import openpyxl
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils.cell import column_index_from_string, coordinate_from_string

from app.core.config import settings
from app.core.constants import COLOR_MAP as _COLOR_MAP, get_pisos_config, get_spt_values
from app.services.resource_manager import resource_manager
from app.utils.common import normalize as _common_normalize, split_project_location as _common_split_location
from app.utils.field_mapping import (
    get_field_mapping,
    get_parametros_mapping,
    get_perforacion_mapping,
)

logger = logging.getLogger(__name__)

MAIN_NS = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"
REL_NS = "http://schemas.openxmlformats.org/officeDocument/2006/relationships"
XML_NS = "http://www.w3.org/XML/1998/namespace"
NS = {"main": MAIN_NS, "rel": REL_NS}

# Pre-compiled regex for AlternateContent removal — avoids recompilation on every call
_RE_ALTERNATE_CONTENT = re.compile(
    r'<\/?(?:\w+:)?AlternateContent[^>]*>(?:.|\n|\r)*?<\/?(?:\w+:)?AlternateContent>'
)


class ExcelService:
    """Service responsible for copying and filling Excel templates."""

    def __init__(self):
        self.templates_dir = settings.EXCEL_TEMPLATES_DIR
        self.generated_dir = settings.GENERATED_DIR

    def _normalize(self, value: str) -> str:
        return unicodedata.normalize("NFD", value).encode("ascii", "ignore").decode("ascii").lower()

    def _get_template_path(self, template_id: str) -> Path:
        template_filename = settings.TEMPLATES_CONFIG.get(str(template_id))
        if not template_filename:
            raise FileNotFoundError(f"Plantilla no configurada para template_id={template_id}")

        template_path = self.templates_dir / template_filename
        if not template_path.exists():
            raise FileNotFoundError(f"Plantilla no encontrada: {template_path}")

        return template_path

    def _copy_template(self, template_id: str, project_id: str) -> Path:
        template_path = self._get_template_path(template_id)
        output_filename = f"{project_id}_template_{template_id}{template_path.suffix}"
        output_path = self.generated_dir / output_filename
        shutil.copy2(template_path, output_path)
        logger.info("Plantilla copiada: %s -> %s", template_path, output_path)
        return output_path

    def _calculate_spt_values(self, template_id: str, use_lower: bool, pisos_int: int = 0) -> List[str]:
        """Devuelve los valores SPT para la plantilla y el toggle indicados.

        Delega a constants.get_spt_values — fuente única de verdad para las
        secuencias SPT.  El parámetro pisos_int se conserva por compatibilidad
        pero ya no se usa: el template_id (o su equivalente de correlación) es
        el único indicador del rango de pisos.
        """
        _template_mapping = {"12": "1", "13": "2", "14": "3", "15": "3"}
        effective = _template_mapping.get(str(template_id), str(template_id))
        if effective in ("1", "2", "3"):
            return get_spt_values(effective, use_lower)
        return []

    def _save_spt_state(self, effective_template_id: str, values: List[str]) -> None:
        """Persist SPT values so a paired legacy template can reuse them exactly."""
        try:
            state_file = self.generated_dir / f".spt_state_{effective_template_id}.json"
            state_file.write_text(json.dumps(values))
        except Exception:
            logger.debug("No se pudieron guardar los valores SPT para template %s", effective_template_id, exc_info=True)

    def _load_spt_state(self, effective_template_id: str) -> Optional[List[str]]:
        """Read previously persisted SPT values for a given template."""
        try:
            state_file = self.generated_dir / f".spt_state_{effective_template_id}.json"
            if state_file.exists():
                return json.loads(state_file.read_text())
        except Exception:
            logger.debug("No se pudieron cargar los valores SPT para template %s", effective_template_id, exc_info=True)
        return None

    def _save_lab_state(self, state: dict) -> None:
        """Persist lab data (n_golpes, lp) so INCONFINADO and ASENTAMIENTOS can reuse it."""
        try:
            state_file = self.generated_dir / ".lab_state.json"
            state_file.write_text(json.dumps(state))
        except Exception:
            logger.debug("No se pudo guardar el estado de laboratorio", exc_info=True)

    def _load_lab_state(self) -> Optional[dict]:
        """Read previously persisted lab state."""
        try:
            state_file = self.generated_dir / ".lab_state.json"
            if state_file.exists():
                return json.loads(state_file.read_text())
        except Exception:
            logger.debug("No se pudo cargar el estado de laboratorio", exc_info=True)
        return None

    def _get_legacy_n_campo_values(self, template_id: str, use_lower: bool, pisos_int: int = 0) -> List[tuple]:
        """Devuelve los valores N° CAMPO para las plantillas .xls heredadas (P-1…P-4).

        Todos los archivos P reflejan la CORRELACIÓN GEOTÉCNICA generada en la
        misma solicitud, de modo que su columna T coincide con la columna F de
        dicha plantilla.  La clave de correlación y el conteo esperado se
        obtienen de constants.get_pisos_config — fuente única de verdad.
        """
        cfg = get_pisos_config(pisos_int)
        correlation_key = cfg["template_id"]
        expected_count = cfg["spt_count"]

        # Reutilizar los valores guardados por la plantilla de correlación en
        # esta misma solicitud.  Rechazar estados con conteo incorrecto (stale).
        saved = self._load_spt_state(correlation_key)
        if saved and len(saved) == expected_count:
            return [(v, "General") for v in saved]

        # Fallback: calcular desde constants usando opt1 para coherencia.
        try:
            return [(v, "General") for v in get_spt_values(correlation_key, True)]
        except Exception:
            logger.debug("No se pudieron calcular los valores de N° CAMPO para template %s", template_id, exc_info=True)
            return []

    def _fill_legacy_xls_template(
        self,
        workbook_path: Path,
        project_value,
        fecha_value,
        e5_value,
        n_campo_values: Optional[List[tuple]] = None,
        soil_descriptions: Optional[List[str]] = None,
        soil_colors: Optional[List[str]] = None,
        depth_values: Optional[List[float]] = None,
        expand_depth_levels: Optional[int] = None,
        photo_paths: Optional[List[Path]] = None,
        lab_data_per_layer: Optional[List[Optional[dict]]] = None,
        municipio: Optional[str] = None,
    ) -> None:
        try:
            import pythoncom
            import win32com.client
        except ImportError as exc:
            # If pywin32 is not available, log a warning and skip editing the
            # legacy .xls file so the generation flow can continue and the
            # unmodified template copy will still be included in the ZIP. The
            # frontend will receive a friendly message if editing is required
            # for the .xls files to be pre-populated.
            logger.warning("pywin32 no instalado: se omitirá edición de plantilla .xls (%s)", workbook_path)
            return

        excel_app = None
        workbook = None
        pythoncom.CoInitialize()
        try:
            excel_app = win32com.client.DispatchEx("Excel.Application")
            excel_app.Visible = False
            excel_app.DisplayAlerts = False
            workbook = excel_app.Workbooks.Open(str(workbook_path))
            worksheet = workbook.Worksheets(1)

            project_text = str(project_value or "").strip().upper()
            if project_text:
                worksheet.Range("B3").Value = project_text

            if fecha_value is not None:
                if isinstance(fecha_value, date) and not isinstance(fecha_value, datetime):
                    from datetime import datetime as _dt
                    fecha_value = _dt.combine(fecha_value, _dt.min.time())
                elif isinstance(fecha_value, datetime):
                    fecha_value = fecha_value
                else:
                    try:
                        fecha_value = datetime.fromisoformat(str(fecha_value))
                    except Exception:
                        fecha_value = str(fecha_value)
                worksheet.Range("N3").NumberFormat = "yyyy-mm-dd"
                worksheet.Range("N3").Value = fecha_value

            if e5_value is not None:
                worksheet.Range("E5").Value = e5_value

            if municipio:
                worksheet.Range("O5").Value = municipio

            # Update column B with SPT sample depths (0.45, 1.45, 2.45…).
            # Slots: B8=level0, B10=level1, B12=level2, B14=level3, B16=level4, B18=level5, B20=level6.
            # Skipped when expand_depth_levels is set (B handled by the scale block below).
            if expand_depth_levels is None:
                _b_all_slots = [8, 10, 12, 14, 16, 18, 20]
                for _b_idx, _b_row in enumerate(_b_all_slots):
                    try:
                        worksheet.Range(f"B{_b_row}").Value = round(_b_idx + 0.45, 2)
                    except Exception:
                        pass

            # Dynamic depth scale expansion for P-1 (and similar single-column-B templates).
            # The template has 7 levels (0-6) in rows 8-21, each occupying 2 merged rows.
            # Rows are inserted before OBSERVACIONES (row 22) when more levels are required.
            # Charts at D7/F7 and images in C8:C19 are above the insertion point and unaffected.
            if expand_depth_levels is not None:
                _B_FIRST_ROW = 8      # B8 = depth 0 in the template
                _TEMPLATE_LEVELS = 7  # template contains levels 0-6
                _ROWS_PER = 2         # each level occupies 2 merged rows
                _OBS_START = _B_FIRST_ROW + _TEMPLATE_LEVELS * _ROWS_PER  # = 22

                target_levels = int(expand_depth_levels)

                if target_levels > _TEMPLATE_LEVELS:
                    extra_levels = target_levels - _TEMPLATE_LEVELS
                    extra_rows = extra_levels * _ROWS_PER
                    ins_end = _OBS_START + extra_rows - 1

                    # Insert rows shifting OBSERVACIONES block down.
                    worksheet.Rows(f"{_OBS_START}:{ins_end}").Insert()

                    # Merge B and K for each new level and apply basic alignment.
                    for _lvl in range(_TEMPLATE_LEVELS, target_levels):
                        _r1 = _B_FIRST_ROW + _lvl * _ROWS_PER
                        _r2 = _r1 + 1

                        worksheet.Rows(_r1).RowHeight = 15.0
                        worksheet.Rows(_r2).RowHeight = 15.0

                        _b = worksheet.Range(f"B{_r1}:B{_r2}")
                        _b.Merge()
                        _b.HorizontalAlignment = -4108  # xlCenter
                        _b.VerticalAlignment = -4108    # xlCenter

                        _k = worksheet.Range(f"K{_r1}:K{_r2}")
                        _k.Merge()
                        _k.HorizontalAlignment = -4108
                        _k.VerticalAlignment = -4108

                # Write SPT sample depths (0.45, 1.45, 2.45…) to column B.
                # These are the midpoint depths for each SPT test interval.
                for _lvl in range(target_levels):
                    _row = _B_FIRST_ROW + _lvl * _ROWS_PER
                    worksheet.Range(f"B{_row}").Value = round(_lvl + 0.45, 2)

            # Write N campo values to column T starting at row 13.
            # MUST run after expand_depth_levels inserts rows — if written before,
            # the row insertion at row 22 shifts T22+ down and splits the sequence.
            if n_campo_values:
                _t_count = len(n_campo_values)
                # Clear T and U from row 13 to a safe upper bound to remove any
                # stale template default values and rows shifted by row expansion.
                _clear_end = max(30, 12 + _t_count + 5)
                for _tr in range(13, _clear_end + 1):
                    try:
                        worksheet.Range(f"T{_tr}").ClearContents()
                    except Exception:
                        pass
                    try:
                        worksheet.Range(f"U{_tr}").ClearContents()
                    except Exception:
                        pass

                for offset, (value, number_format) in enumerate(n_campo_values, start=13):
                    i = offset - 13  # 0-indexed level
                    cell = worksheet.Range(f"T{offset}")
                    cell.Value = value
                    if number_format:
                        cell.NumberFormat = number_format
                    # Write Z(m) depth for this level: 0.45, 1.45, 2.45, ...
                    u_cell = worksheet.Range(f"U{offset}")
                    u_cell.Value = round(i + 0.45, 2)
                    u_cell.NumberFormat = "0.00"

                # Write K column formulas =T13, =T14, ... down to the last B row.
                # K column uses 2 rows per level (merged): K8, K10, K12, ...
                # Each K cell references the matching T row so N CAMPO in the depth
                # scale area stays in sync with the T summary table.
                _K_FIRST = 8
                _K_STEP = 2
                for _ki in range(_t_count):
                    _k_row = _K_FIRST + _ki * _K_STEP
                    _t_ref = 13 + _ki
                    try:
                        worksheet.Range(f"K{_k_row}").Formula = f"=T{_t_ref}"
                    except Exception:
                        pass

            # === PROPORTIONAL COLUMN I — Descripción Macroscópica ===
            # Uses _calc_layer_spans (same algorithm as column Q of the correlation .xlsx)
            # so strata proportions are always identical between both documents.
            if soil_descriptions is not None:
                _I_FIRST = 8
                _ROWS_PER_LVL = 2

                if expand_depth_levels is not None:
                    _i_total_rows = int(expand_depth_levels) * _ROWS_PER_LVL
                else:
                    _i_total_rows = 7 * _ROWS_PER_LVL

                _i_last = _I_FIRST + _i_total_rows - 1

                # Clear and unmerge the I:J zone
                _zone = worksheet.Range(f"I{_I_FIRST}:J{_i_last}")
                _zone.UnMerge()
                _zone.ClearContents()
                _zone.Interior.ColorIndex = -4142  # xlColorIndexNone
                _zone.Font.Color = 0
                _zone.Font.Bold = False

                # Build layer dicts compatible with _calc_layer_spans
                _descs = list(soil_descriptions)
                _colors = list(soil_colors) if soil_colors else []
                _depths = list(depth_values) if depth_values else []
                _n = max(len(_descs), len(_colors), len(_depths))
                _descs += [""] * (_n - len(_descs))
                _colors += [""] * (_n - len(_colors))
                _depths += [0.0] * (_n - len(_depths))

                _layer_dicts = [
                    {"descripcion_suelo": d, "color_predominante": c, "profundidad_z": z}
                    for d, c, z in zip(_descs, _colors, _depths)
                    if d or c
                ]

                _layer_spans = self._calc_layer_spans(_layer_dicts, _i_total_rows)

                # Coalesce adjacent segments with same description (mirrors column Q logic)
                _coalesced: List[list] = []
                for _lyr, _sp in _layer_spans:
                    _dk = _lyr.get("descripcion_suelo", "")
                    if _coalesced and _coalesced[-1][0].get("descripcion_suelo", "") == _dk:
                        _coalesced[-1][1] += _sp
                    else:
                        _coalesced.append([_lyr, _sp])
                _layer_spans = [tuple(x) for x in _coalesced]

                _xl_center = -4108  # xlCenter
                _cur = _I_FIRST
                for _lyr, _sp in _layer_spans:
                    _end = min(_i_last, _cur + _sp - 1)
                    _d = _lyr.get("descripcion_suelo", "")
                    _c = _lyr.get("color_predominante", "")
                    _rng = worksheet.Range(f"I{_cur}:J{_end}")
                    _rng.Merge()
                    if _d:
                        _rng.Value = str(_d)
                    _rng.WrapText = True
                    _rng.HorizontalAlignment = _xl_center
                    _rng.VerticalAlignment = _xl_center
                    _rng.Font.Size = 9
                    if _c:
                        _fh, _fnt = self._get_color_info(_c)
                        _rng.Interior.Color = self._hex_to_xls_color(_fh)
                        _rng.Font.Color = self._hex_to_xls_color(_fnt)
                    _cur = _end + 1
                    if _cur > _i_last:
                        break

            # === COLUMN C — Replace or redistribute soil-sketch images/shapes ===
            # If photo_paths is provided, delete the template sketches and insert
            # the user's field photos in their place, distributed across the data
            # area.  Otherwise fall back to the original redistribution logic.
            try:
                _c_n_levels = int(expand_depth_levels) if expand_depth_levels is not None else 7
                _c_first_row = 8
                _c_rows_per = 2
                _c_area_top = worksheet.Rows(_c_first_row).Top
                _c_last_row_num = _c_first_row + _c_n_levels * _c_rows_per - 1
                _c_last_row_obj = worksheet.Rows(_c_last_row_num)
                _c_area_bottom = _c_last_row_obj.Top + _c_last_row_obj.Height
                _c_total_h = _c_area_bottom - _c_area_top

                _col_c = worksheet.Columns("C")
                _col_c_left = _col_c.Left
                _col_c_width = _col_c.Width
                _col_c_right = _col_c_left + _col_c_width

                # Collect existing column-C shapes so we can delete or reuse them.
                _c_shapes = []
                for _sh in worksheet.Shapes:
                    try:
                        if not _sh.Visible:
                            continue
                        _sh_cx = _sh.Left + _sh.Width / 2
                        if _col_c_left <= _sh_cx <= _col_c_right:
                            if _sh.Top >= _c_area_top - 5:
                                _c_shapes.append(_sh)
                    except Exception:
                        pass
                _c_shapes.sort(key=lambda s: s.Top)

                if photo_paths and _c_total_h > 0:
                    # Delete every existing column-C sketch.
                    for _sh in _c_shapes:
                        try:
                            _sh.Delete()
                        except Exception:
                            pass

                    # Insert user photos distributed evenly across the data area.
                    _slot_h = _c_total_h / len(photo_paths)
                    for _idx, _photo in enumerate(photo_paths):
                        try:
                            _top = _c_area_top + _idx * _slot_h
                            worksheet.Shapes.AddPicture(
                                str(_photo.resolve()),  # Filename (absolute path)
                                0,                      # LinkToFile = msoFalse
                                -1,                     # SaveWithDocument = msoTrue
                                _col_c_left,            # Left
                                _top,                   # Top
                                _col_c_width,           # Width
                                _slot_h,                # Height
                            )
                        except Exception:
                            logger.debug(
                                "No se pudo insertar foto %s en columna C de %s",
                                _photo, workbook_path, exc_info=True,
                            )
                elif _c_shapes and _c_total_h > 0:
                    # Fallback: redistribute existing template sketches.
                    _slot_h = _c_total_h / len(_c_shapes)
                    for _idx, _sh in enumerate(_c_shapes):
                        try:
                            _sh.LockAspectRatio = 0   # msoFalse — allow free resize
                            _sh.Top = _c_area_top + _idx * _slot_h
                            _sh.Height = _slot_h
                            _sh.Left = _col_c_left
                            _sh.Width = _col_c_width
                        except Exception:
                            pass
            except Exception:
                logger.debug(
                    "No se pudieron redistribuir las imágenes de la columna C en %s",
                    workbook_path, exc_info=True,
                )

            # === EXPAND CHARTS down to the last row with data in column B ===
            try:
                _cht_n_levels = int(expand_depth_levels) if expand_depth_levels is not None else 7
                _cht_first_row = 8
                _cht_rows_per = 2
                _cht_last_row_num = _cht_first_row + _cht_n_levels * _cht_rows_per - 1
                _cht_last_row_obj = worksheet.Rows(_cht_last_row_num)
                _cht_area_bottom = _cht_last_row_obj.Top + _cht_last_row_obj.Height

                _cht_col = worksheet.ChartObjects()
                for _ci in range(1, _cht_col.Count + 1):
                    try:
                        _co = _cht_col(_ci)
                        _new_h = _cht_area_bottom - _co.Top
                        if _new_h > 0:
                            _co.Height = _new_h
                    except Exception:
                        pass
            except Exception:
                logger.debug(
                    "No se pudieron expandir las gráficas en %s",
                    workbook_path, exc_info=True,
                )

            # === LAB RESULTS — columns L (w), M (LL), N (LP), O (IP), P (γ), R (USCS) ===
            # One row per layer: row = 8 + layer_index * 2
            # First clear all template placeholder values in these columns (rows 8-20).
            if lab_data_per_layer:
                try:
                    _LAB_FIRST_ROW = 8
                    _LAB_ROWS_PER = 2
                    _LAB_COLS = ('L', 'M', 'N', 'O', 'P', 'Q', 'R', 'S')
                    _n_levels = int(expand_depth_levels) if expand_depth_levels is not None else 7
                    for _ci in range(_n_levels):
                        _cr = _LAB_FIRST_ROW + _ci * _LAB_ROWS_PER
                        for _col in _LAB_COLS:
                            try:
                                worksheet.Range(f"{_col}{_cr}").ClearContents()
                            except Exception:
                                pass

                    for _li, _ldata in enumerate(lab_data_per_layer):
                        if _ldata is None:
                            continue
                        _lr = _LAB_FIRST_ROW + _li * _LAB_ROWS_PER
                        _w   = _ldata.get('humedad')
                        _ll  = _ldata.get('limite_liquido')
                        _lp  = _ldata.get('limite_plastico')
                        _ip  = _ldata.get('indice_plasticidad')
                        _gam = _ldata.get('gamma')
                        _cls = _ldata.get('clasificacion_uscs')
                        if _w is not None:
                            worksheet.Range(f"L{_lr}").Value = _w
                            worksheet.Range(f"L{_lr}").NumberFormat = "0.0"
                        if _ll is not None:
                            worksheet.Range(f"M{_lr}").Value = _ll
                            worksheet.Range(f"M{_lr}").NumberFormat = "0"
                        if _lp is not None:
                            worksheet.Range(f"N{_lr}").Value = _lp
                            worksheet.Range(f"N{_lr}").NumberFormat = "0"
                        if _ip is not None:
                            worksheet.Range(f"O{_lr}").Value = _ip
                            worksheet.Range(f"O{_lr}").NumberFormat = "0"
                        if _gam is not None:
                            worksheet.Range(f"P{_lr}").Value = _gam
                            worksheet.Range(f"P{_lr}").NumberFormat = "0.0"
                        if _cls:
                            worksheet.Range(f"R{_lr}").Value = str(_cls).upper()
                except Exception:
                    logger.debug("No se pudieron escribir datos de laboratorio en %s", workbook_path, exc_info=True)

            workbook.Save()
        finally:
            if workbook is not None:
                try:
                    workbook.Close(SaveChanges=False)
                except Exception:
                    logger.debug("No se pudo cerrar el libro .xls de forma limpia", exc_info=True)
            if excel_app is not None:
                try:
                    excel_app.Quit()
                except Exception:
                    logger.debug("No se pudo cerrar la instancia de Excel", exc_info=True)
            pythoncom.CoUninitialize()

    def _set_cell_value(self, worksheet, cell_ref: str, value):
        if ":" in cell_ref:
            cell_ref = cell_ref.split(":")[0]

        cell = worksheet[cell_ref]
        if cell.__class__.__name__ == "MergedCell":
            column_letter, row_number = coordinate_from_string(cell_ref)
            column_number = column_index_from_string(column_letter)
            for merged_range in worksheet.merged_cells.ranges:
                if merged_range.min_row <= row_number <= merged_range.max_row and merged_range.min_col <= column_number <= merged_range.max_col:
                    worksheet[merged_range.start_cell.coordinate] = value
                    return

        worksheet[cell_ref] = value

    def _get_sheet(self, workbook, target_name: str):
        normalized_target = self._normalize(target_name)
        for sheet_name in workbook.sheetnames:
            if self._normalize(sheet_name) == normalized_target:
                return workbook[sheet_name]
        return workbook.active

    def _resolve_sheet_targets(self, workbook_path: Path) -> dict:
        with ZipFile(workbook_path, "r") as workbook_zip:
            workbook_xml = ET.fromstring(workbook_zip.read("xl/workbook.xml"))
            rels_xml = ET.fromstring(workbook_zip.read("xl/_rels/workbook.xml.rels"))

        rel_map = {rel.get("Id"): rel.get("Target") for rel in rels_xml}
        sheet_targets = {}

        for sheet in workbook_xml.findall("main:sheets/main:sheet", NS):
            sheet_name = sheet.get("name", "")
            rel_id = sheet.get(f"{{{REL_NS}}}id")
            target = rel_map.get(rel_id)
            if not target:
                continue

            if target.startswith("/"):
                target = target.lstrip("/")
            if not target.startswith("xl/"):
                target = f"xl/{target}"

            sheet_targets[self._normalize(sheet_name)] = target

        return sheet_targets

    def _cell_sort_key(self, cell_ref: str):
        letters = ""
        digits = ""
        for character in cell_ref:
            if character.isdigit():
                digits += character
            else:
                letters += character
        return int(digits or 0), column_index_from_string(letters or "A")

    # Mapa de colores importado desde constants.py — fuente única de verdad.
    # Claves pre-normalizadas (sin tildes, minúsculas).  Usado por las rutas
    # openpyxl (.xlsx) y win32com (.xls).  NO redefinir aquí.

    def _get_color_info(self, color_name: Optional[str]) -> tuple:
        """Devuelve (fill_hex_RRGGBB, font_hex_RRGGBB) para un nombre de color.

        Usa _COLOR_MAP (constants.py) con claves pre-normalizadas y búsqueda
        directa O(1) mediante _common_normalize — fuente única de verdad.
        """
        normalized = _common_normalize(color_name or "")
        fill_hex = _COLOR_MAP.get(normalized, "F4B183")
        try:
            r = int(fill_hex[0:2], 16)
            g = int(fill_hex[2:4], 16)
            b = int(fill_hex[4:6], 16)
            is_dark = (r * 299 + g * 587 + b * 114) / 1000 < 140
        except Exception:
            is_dark = False
        font_hex = "FFFFFF" if is_dark else "1F2937"
        return fill_hex, font_hex

    @staticmethod
    def _hex_to_xls_color(hex_rgb: str) -> int:
        """Convert RRGGBB hex string to win32com Interior.Color integer (R+G*256+B*65536)."""
        r = int(hex_rgb[0:2], 16)
        g = int(hex_rgb[2:4], 16)
        b = int(hex_rgb[4:6], 16)
        return r + g * 256 + b * 65536

    def _soil_style_from_color(self, color_name: Optional[str]):
        fill_hex, font_color = self._get_color_info(color_name)
        return PatternFill(fill_type="solid", fgColor=fill_hex), font_color

    def _parse_depth_value(self, value) -> Optional[float]:
        try:
            if value in (None, ""):
                return None
            parsed = float(str(value).replace(",", ".").strip())
            return parsed if parsed > 0 else None
        except Exception:
            return None

    def _clean_soil_text(self, value: Optional[str], color_name: Optional[str] = None) -> str:
        text = str(value or "").strip()
        if not text:
            return ""

        normalized = self._normalize(text)
        for separator in (" de color ", " color "):
            token = separator.strip()
            index = normalized.find(token)
            if index != -1:
                text = text[:index].strip()
                normalized = self._normalize(text)

        if color_name:
            normalized_color = self._normalize(color_name)
            if normalized_color:
                index = normalized.find(f" color {normalized_color}")
                if index != -1:
                    text = text[:index].strip()

        text = text.replace("de color", "").replace(" color ", " ")
        return " ".join(text.split())

    def _calc_layer_spans(self, layers: List[dict], total_rows: int) -> List[tuple]:
        """Proportional row-span calculator — single source of truth for strata layout.

        Both column Q of the correlation (.xlsx) and column I of P-1/P-2/P-3/P-4
        call this method so their stratum proportions are always identical.

        Returns a list of (layer_dict, num_rows) pairs whose spans sum to total_rows.
        """
        valid_layers = [
            l for l in layers
            if any(l.get(k) for k in ("descripcion_suelo", "tipo_suelo_principal", "color_predominante"))
        ]
        if not valid_layers:
            return []

        end_depths: List[float] = []
        for layer in valid_layers:
            try:
                d = float(layer.get("profundidad_z") or 0)
            except Exception:
                d = 0.0
            end_depths.append(d)

        # Equal distribution fallback when no depths are provided
        if all(d <= 0 for d in end_depths):
            rows_per = max(1, total_rows // len(valid_layers))
            result: List[tuple] = []
            remaining = total_rows
            for i, layer in enumerate(valid_layers):
                if i == len(valid_layers) - 1:
                    result.append((layer, max(1, remaining)))
                else:
                    span = max(1, min(rows_per, remaining - (len(valid_layers) - i - 1)))
                    result.append((layer, span))
                    remaining -= span
            return result

        # Proportional distribution based on profundidad_z
        valid_depths = [d for d in end_depths if d > 0]
        total_depth = max(valid_depths) if valid_depths else float(len(valid_layers))

        raw_spans: List[list] = []
        prev_depth = 0.0
        for layer, end_d in zip(valid_layers, end_depths):
            eff_d = end_d if end_d > 0 else (prev_depth + total_depth / len(valid_layers))
            delta = max(0.0, eff_d - prev_depth)
            span = max(1, int(round(delta / total_depth * total_rows)))
            raw_spans.append([layer, span])
            prev_depth = eff_d

        # Adjust last span to fill exactly total_rows
        diff = total_rows - sum(s for _, s in raw_spans)
        if raw_spans:
            raw_spans[-1][1] = max(1, raw_spans[-1][1] + diff)
            while sum(s for _, s in raw_spans) > total_rows and raw_spans[-1][1] > 1:
                raw_spans[-1][1] -= 1

        return [(layer, span) for layer, span in raw_spans]

    def _build_soil_segments(self, worksheet, layers: List[dict], start_row: int, end_row: int) -> List[tuple]:
        """Convert layer list into (row_start, row_end, layer) segments.

        Delegates to _calc_layer_spans so column Q proportions match column I
        of P-1/P-2/P-3/P-4 exactly. The worksheet parameter is accepted for
        signature compatibility but is no longer read.
        """
        total_rows = end_row - start_row + 1
        layer_spans = self._calc_layer_spans(layers, total_rows)

        segments: List[tuple] = []
        cur = start_row
        for layer, span in layer_spans:
            seg_end = min(end_row, cur + span - 1)
            segments.append((cur, seg_end, layer))
            cur = seg_end + 1
            if cur > end_row:
                break

        # Ensure last segment always reaches end_row
        if segments:
            last = segments[-1]
            segments[-1] = (last[0], end_row, last[2])

        return segments

    def _build_depth_soil_segments_exact(
        self,
        layers: List[dict],
        start_row: int,
        end_row: int,
    ) -> List[tuple]:
        """Build Q segments by matching fixed row depths to user-defined layer end-depths.

        Row depths are 0.45, 1.45, 2.45, … starting at start_row.
        For each row depth D:
          - Find the FIRST layer whose profundidad_z >= D.  That is the layer whose
            depth range [prev_end, current_end] contains D.
          - If D exceeds the deepest layer's end-depth → extend the last layer.
        Accepts comma or dot as decimal separator in profundidad_z.
        """
        def _to_float(v) -> float:
            try:
                return float(str(v or 0).replace(',', '.'))
            except Exception:
                return 0.0

        valid_layers = [
            layer for layer in layers
            if any(layer.get(k) for k in ("descripcion_suelo", "tipo_suelo_principal", "color_predominante"))
        ]
        if not valid_layers:
            return []

        sorted_layers = sorted(
            valid_layers,
            key=lambda l: (_to_float(l.get('profundidad_z')) or float('inf')),
        )
        depths = [_to_float(l.get('profundidad_z')) for l in sorted_layers]
        valid_depths = [d for d in depths if d > 0]
        last_depth = max(valid_depths) if valid_depths else 0.0
        if last_depth == 0.0:
            return []

        row_count = end_row - start_row + 1
        last_layer_idx = len(sorted_layers) - 1
        row_assignments: List = []
        for i in range(row_count):
            row_depth = round(0.45 + i * 1.0, 10)
            if row_depth > last_depth + 1e-9:
                # Extend the last defined layer to fill all remaining column A rows.
                row_assignments.append(last_layer_idx)
            else:
                # Find the FIRST layer whose end-depth >= row_depth.
                # A row at depth D belongs to the layer whose range contains D,
                # i.e. the layer that ends at or after D (first such layer in sorted order).
                active_idx = last_layer_idx  # fallback: last layer
                for j, d in enumerate(depths):
                    if d > 0 and d >= row_depth - 1e-9:
                        active_idx = j
                        break
                row_assignments.append(active_idx)

        segments: List[tuple] = []
        if not row_assignments:
            return []

        seg_idx = row_assignments[0]
        seg_start = start_row

        for i in range(1, len(row_assignments)):
            cur_idx = row_assignments[i]
            if cur_idx == seg_idx:
                continue
            if seg_idx is not None:
                segments.append((seg_start, start_row + i - 1, sorted_layers[seg_idx]))
            seg_idx = cur_idx
            seg_start = start_row + i

        if seg_idx is not None:
            segments.append((seg_start, end_row, sorted_layers[seg_idx]))

        return segments

    def _apply_q_segment_format(
        self,
        sheet_obj,
        row_start: int,
        row_end: int,
        layer: dict,
        soil_text: str,
    ) -> None:
        """Apply visual format to one Q column segment.

        Steps:
          1. Fill EVERY cell in [row_start, row_end] with the layer colour so the
             entire combined region has a uniform background (not just the anchor).
          2. Merge the range when it spans more than one row.
          3. Write the soil text once on the anchor (top-left cell), centred both
             horizontally and vertically, with word-wrap enabled.

        This is the single reusable function used for all three correlation
        templates (1, 2, 3) so no formatting logic is duplicated.
        """
        fill, font_color = self._soil_style_from_color(layer.get('color_predominante'))

        # Step 1 — colour every individual cell BEFORE merging
        for _r in range(row_start, row_end + 1):
            try:
                sheet_obj.cell(row=_r, column=17).fill = fill
            except Exception:
                pass

        # Step 2 — merge when multi-row
        if row_start != row_end:
            try:
                sheet_obj.merge_cells(f"Q{row_start}:Q{row_end}")
            except Exception:
                pass

        # Step 3 — text and formatting on the anchor only
        q_anchor = sheet_obj[f'Q{row_start}']
        q_anchor.value = soil_text
        q_anchor.fill = fill  # re-confirm after merge (openpyxl may reset anchor style)
        q_anchor.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        soil_font = copy(q_anchor.font)
        soil_font.size = 7  # normalise across all templates
        soil_font.color = font_color
        q_anchor.font = soil_font

    def _clear_cell_children(self, cell):
        for child in list(cell):
            cell.remove(child)

    def _set_xml_cell_value(self, cell, value):
        self._clear_cell_children(cell)

        if value is None or value == "":
            return

        if isinstance(value, bool):
            cell.set("t", "b")
            v_element = ET.SubElement(cell, f"{{{MAIN_NS}}}v")
            v_element.text = "1" if value else "0"
            return

        if isinstance(value, (int, float)):
            cell.set("t", "n")
            v_element = ET.SubElement(cell, f"{{{MAIN_NS}}}v")
            v_element.text = str(value)
            return

        cell.set("t", "inlineStr")
        inline_string = ET.SubElement(cell, f"{{{MAIN_NS}}}is")
        text_element = ET.SubElement(inline_string, f"{{{MAIN_NS}}}t")
        text_value = str(value)
        if text_value != text_value.strip():
            text_element.set(f"{{{XML_NS}}}space", "preserve")
        text_element.text = text_value

    def _ensure_row(self, sheet_data, row_number: int):
        rows = sheet_data.findall(f"{{{MAIN_NS}}}row")
        for row in rows:
            if int(row.get("r", "0")) == row_number:
                return row

        new_row = ET.Element(f"{{{MAIN_NS}}}row", {"r": str(row_number)})
        insert_at = len(rows)
        for index, row in enumerate(rows):
            if int(row.get("r", "0")) > row_number:
                insert_at = index
                break
        sheet_data.insert(insert_at, new_row)
        return new_row

    def _ensure_cell(self, row, cell_ref: str):
        cells = row.findall(f"{{{MAIN_NS}}}c")
        for cell in cells:
            if cell.get("r") == cell_ref:
                return cell

        new_cell = ET.Element(f"{{{MAIN_NS}}}c", {"r": cell_ref})
        insert_at = len(cells)
        target_sort_key = self._cell_sort_key(cell_ref)
        for index, cell in enumerate(cells):
            if self._cell_sort_key(cell.get("r", "A1")) > target_sort_key:
                insert_at = index
                break
        row.insert(insert_at, new_cell)
        return new_cell

    def _update_sheet_xml(self, sheet_xml: bytes, cell_updates: dict) -> bytes:
        root = ET.fromstring(sheet_xml)
        sheet_data = root.find(f"{{{MAIN_NS}}}sheetData")
        if sheet_data is None:
            return sheet_xml

        for cell_ref, value in cell_updates.items():
            digits = "".join(ch for ch in cell_ref if ch.isdigit())
            row_number = int(digits or "0")
            row = self._ensure_row(sheet_data, row_number)
            cell = self._ensure_cell(row, cell_ref)
            self._set_xml_cell_value(cell, value)

        

        return ET.tostring(root, encoding="utf-8", xml_declaration=True)

    def _write_workbook_preserving_parts(self, workbook_path: Path, sheet_updates: dict) -> None:
        with ZipFile(workbook_path, "r") as source_zip:
            source_entries = [(info, source_zip.read(info.filename)) for info in source_zip.infolist()]

        with NamedTemporaryFile(delete=False, suffix=".xlsx") as temp_file:
            temp_path = Path(temp_file.name)

        try:
            with ZipFile(temp_path, "w", compression=ZIP_DEFLATED) as target_zip:
                for info, data in source_entries:
                    # sanitize workbook.xml to remove AlternateContent blocks that reference absolute paths
                    if info.filename == 'xl/workbook.xml':
                        try:
                            wb_root = ET.fromstring(data)
                            s = ET.tostring(wb_root, encoding='utf-8')
                            try:
                                s_dec = s.decode('utf-8')
                                s_clean = _RE_ALTERNATE_CONTENT.sub('', s_dec)
                                data = s_clean.encode('utf-8')
                            except Exception:
                                pass
                        except Exception:
                            pass
                    if info.filename in sheet_updates:
                        replacement = sheet_updates[info.filename]
                        if isinstance(replacement, (bytes, bytearray)):
                            data = replacement
                        else:
                            # assume mapping of cell updates
                            data = self._update_sheet_xml(data, replacement)
                    target_zip.writestr(info, data)
            # replace original workbook atomically when possible
            try:
                if workbook_path.exists():
                    workbook_path.unlink()
                shutil.move(str(temp_path), str(workbook_path))
            except Exception:
                # best-effort move; re-raise to let caller handle
                raise
        finally:
            if temp_path.exists():
                try:
                    temp_path.unlink()
                except Exception:
                    pass

    def _fill_general_fields(self, worksheet, template_id: str, data: dict):
        field_mapping = get_field_mapping(template_id)
        for field_name, cell_refs in field_mapping.items():
            value = data.get(field_name)
            if value is None:
                continue

            if isinstance(value, date):
                value = value.strftime("%d/%m/%Y")

            for cell_ref in cell_refs:
                self._set_cell_value(worksheet, cell_ref, value)

    def _format_client(self, value: Optional[str]) -> Optional[str]:
        if value is None:
            return None
        try:
            return str(value).strip().upper()
        except Exception:
            return str(value)

    def _split_project_location(self, value: Optional[str]) -> tuple:
        """Delega a common.split_project_location — fuente única de verdad."""
        return _common_split_location(value)

    def _fill_rows(self, worksheet, mapping: dict, rows: List[dict], seed: str = ""):
        if not rows:
            return

        start_row = mapping["tabla_inicio_row"]
        allowed_columns = mapping["allowed_columns"]
        rng = random.SystemRandom()
        reduce_next = bool(rng.getrandbits(1))

        for index, row_data in enumerate(rows):
            row_number = start_row + index
            for field_name, column_letter in allowed_columns.items():
                if field_name == "profundidad_z":
                    continue

                value = row_data.get(field_name, "")

                if field_name == "n_campo_spt" and 10 <= row_number <= 16 and isinstance(value, int):
                    value = max(0, value - 1) if reduce_next else value
                    reduce_next = not reduce_next

                self._set_cell_value(worksheet, f"{column_letter}{row_number}", value)

    # ------------------------------------------------------------------
    # USCS Atterberg helper — single-point Casagrande method (N=25)
    # ------------------------------------------------------------------
    @staticmethod
    def _random_tara_ll() -> float:
        """Random tara in [5.41, 6.69], 2 decimals, second decimal digit never 0."""
        while True:
            v = round(random.uniform(5.41, 6.69), 2)
            if round(v * 100) % 10 != 0:   # rejects *.X0 (would display as 1 decimal)
                return v

    @staticmethod
    def _random_nonzero_decimal(low: float, high: float) -> float:
        """Random float in [low, high], 2 decimals, second decimal digit never 0."""
        while True:
            v = round(random.uniform(low, high), 2)
            if round(v * 100) % 10 != 0:   # rejects *.X0 (would display as 1 decimal)
                return v

    @staticmethod
    def _atterberg_for_clasificacion(clasificacion: str) -> Optional[tuple]:
        """Return (ll, lp) int values that produce the given USCS classification.

        Uses the A-line criterion embedded in the Excel classification formula:
          S27 reads R27 (LL) and F32 (IP) and compares IP vs (LL-20)*0.73.
          Below the A-line → M group;  above → C group.
          LL ≤ 50 → low plasticity (L);  LL > 50 → high plasticity (H).
        """
        cls = (clasificacion or '').strip().upper()
        if not cls:
            return None

        if cls == 'ML':
            # LL ≤ 50, IP below A-line: IP < (LL-20)*0.73
            ll = random.randint(36, 47)
            a_line = (ll - 20) * 0.73
            ip = random.randint(max(5, int(a_line) - 6), max(5, int(a_line) - 1))
            return ll, ll - ip

        if cls == 'MH':
            # LL > 50, IP below A-line
            ll = random.randint(52, 64)
            a_line = (ll - 20) * 0.73
            ip = random.randint(max(5, int(a_line) - 8), max(5, int(a_line) - 1))
            return ll, ll - ip

        if cls == 'CL':
            # LL ≤ 50, IP above A-line (IP ≥ (LL-20)*0.73), LP < 24
            ll = random.randint(33, 46)
            a_line = (ll - 20) * 0.73
            # ip_min ensures: above A-line (ip > a_line) AND lp = ll-ip ≤ 23 (ip ≥ ll-23)
            ip_min = max(int(a_line) + 1, ll - 23)
            ip = random.randint(ip_min, ip_min + 6)
            return ll, max(5, ll - ip)

        if cls == 'CH':
            # LL > 50, IP above A-line
            ll = random.randint(56, 74)
            a_line = (ll - 20) * 0.73
            ip = random.randint(int(a_line) + 1, int(a_line) + 12)
            return ll, max(5, ll - ip)

        if cls in ('SM', 'C'):
            # Very low plasticity: IP < 4 (non-plastic or silty sand)
            ll = random.randint(24, 32)
            ip = random.randint(1, 3)
            return ll, ll - ip

        return None

    def _apply_atterberg_to_sheet(self, target_sheet, ll: float, lp: float):
        """Write LL and LP input cells for a 3-point Casagrande test with realistic weights.

        Returns (n_k, lp_raw) where lp_raw = [[wet+tara, dry+tara, tara], ...]
        for the first 2 LP replicates — used by INCONFINADO for consistency.
        """
        n_k = random.choice([14, 15, 16])
        n_l = random.choice([24, 25, 26])
        n_m = random.choice([34, 35, 36])
        self._set_cell_value(target_sheet, 'K14', n_k)
        self._set_cell_value(target_sheet, 'L14', n_l)
        self._set_cell_value(target_sheet, 'M14', n_m)

        # LL weights — 3 Casagrande points.
        # moisture_N = LL × (25/N)^0.121 so every point lies on the exact
        # Casagrande flow curve and Excel's formula reproduces LL correctly.
        # K16 (dry+tara) ≤ 25, K15 (wet+tara) < 70
        for n, (wet_ref, dry_ref, tara_ref) in [
            (n_k, ('K15', 'K16', 'K17')),
            (n_l, ('L15', 'L16', 'L17')),
            (n_m, ('M15', 'M16', 'M17')),
        ]:
            moisture = ll * (25 / n) ** 0.121
            tara = self._random_tara_ll()
            dry_t = self._random_nonzero_decimal(tara + 9.0, min(24.9, tara + 18.0))
            wet_t = round(dry_t + (dry_t - tara) * moisture / 100, 2)
            self._set_cell_value(target_sheet, tara_ref, tara)
            self._set_cell_value(target_sheet, dry_ref, dry_t)
            self._set_cell_value(target_sheet, wet_ref, wet_t)

        # LP weights — 3 replicates; K23 (dry+tara) ≤ 25, K22 (wet+tara) < 70
        lp_raw = []
        for wet_r, dry_r, tara_r in [
            ('K22', 'K23', 'K24'),
            ('L22', 'L23', 'L24'),
            ('M22', 'M23', 'M24'),
        ]:
            tara = self._random_tara_ll()
            dry_t = self._random_nonzero_decimal(tara + 5.0, min(24.9, tara + 14.0))
            wet_t = round(dry_t + (dry_t - tara) * lp / 100, 2)
            self._set_cell_value(target_sheet, tara_r, tara)
            self._set_cell_value(target_sheet, dry_r, dry_t)
            self._set_cell_value(target_sheet, wet_r, wet_t)
            lp_raw.append([wet_t, dry_t, tara])

        return n_k, lp_raw[:2]

    # G25 base values read data_only from each lab template.
    # Each template has a different Q6 (initial sample weight), which drives
    # E10 → F26 → G26 → X23.  These constants let us compute E26 dynamically
    # so that G26 ≈ target regardless of which template is being filled.
    _SM_G25 = {'4': 85.1635, '5': 82.9811, '6': 80.5239, '7': 80.5239}

    def _set_sm_gradation(self, target_sheet, template_id: str = '4') -> None:
        """Adjust E26 so X23 (% passing #200) lands in the SM range (12 < X23 < 50).

        Chain: Q6 (initial weight) → E10=Q6 → F26=(100/E10)*E26 → G26=G25-F26
               → X23=ROUND(G26,0).
        Each template has its own Q6 (92.34 / 89.9 / 87.8), so E26 must be
        computed from the actual Q6 and the known G25 baseline to hit a target G26.
        """
        try:
            q6 = float(target_sheet['Q6'].value or 92.34)
        except (TypeError, ValueError):
            q6 = 92.34
        g25 = self._SM_G25.get(str(template_id), 85.1635)
        # Random target G26 in [22, 42] → X23 in [22, 42], well inside (12, 50)
        target_g26 = random.uniform(22.0, 42.0)
        e26 = round((g25 - target_g26) * q6 / 100.0, 2)
        self._set_cell_value(target_sheet, 'E26', e26)

    def generate_excel(
        self,
        template_id: str,
        project_id: str,
        data: dict,
        perforaciones: Optional[List[dict]] = None,
        parametros: Optional[List[dict]] = None,
        photo_paths: Optional[List[Path]] = None,
        clasificacion_suelo: Optional[str] = None,
        valores_laboratorio: Optional[dict] = None,
        capacidad_portante: Optional[dict] = None,
        valores_laboratorio_por_lab: Optional[dict] = None,
        clasificaciones_por_lab: Optional[dict] = None,
    ) -> Path:
        work_file = self._copy_template(template_id, project_id)

        # Calculate toggle state once for templates that use SPT values (1, 2, 3, 12-15)
        use_lower = False
        templates_with_spt = {'1', '2', '3', '12', '13', '14', '15'}
        if str(template_id) in templates_with_spt:
            try:
                toggle_file = Path(self.generated_dir) / '.n_campo_toggle'
                if not toggle_file.exists():
                    toggle_file.write_text('0')
                counter = int(toggle_file.read_text() or '0')
                use_lower = (counter % 2 == 0)
                toggle_file.write_text(str(counter + 1))
            except Exception:
                use_lower = False

        try:
            pisos_int = int(data.get('pisos') or 0)
        except Exception:
            pisos_int = 0

        if str(template_id) == '8':
            try:
                wb_tmp = load_workbook(work_file)
                target_sheet = self._get_sheet(wb_tmp, 'MÉTODO TERZAGHI')

                project_value = data.get('proyecto_ubicacion')
                client_value = self._format_client(data.get('cliente'))
                fecha_value = data.get('fecha_registro')

                # Construct header describing the structure levels based on 'pisos'
                pisos_value = data.get('pisos')
                try:
                    pisos_int = int(pisos_value or 0)
                except Exception:
                    pisos_int = 0
                if pisos_int == 1:
                    estructura_text = 'ESTRUCTURA DE CONSTRUCCIÓN DE 1 NIVEL'
                elif pisos_int > 1:
                    estructura_text = f'ESTRUCTURA DE CONSTRUCCIÓN DE {pisos_int} NIVELES'
                else:
                    estructura_text = 'ESTRUCTURA DE CONSTRUCCIÓN'
                self._set_cell_value(target_sheet, 'E1', estructura_text)
                self._set_cell_value(target_sheet, 'E2', project_value)
                self._set_cell_value(target_sheet, 'E3', client_value)
                self._set_cell_value(target_sheet, 'E5', fecha_value)

                # ── Auto-derivar parámetros de capacidad portante ─────────────
                # Prioridad: capacidad_portante del usuario > auto-calculado

                # 1. Profundidad de desplante Df según número de pisos
                if pisos_int <= 1:
                    _df_auto = 1.0
                elif pisos_int <= 3:
                    _df_auto = 1.5
                elif pisos_int <= 6:
                    _df_auto = 2.0
                else:
                    _df_auto = 2.5

                # 2. Tipo de suelo desde clasificacion_suelo (USCS)
                _cls_up8 = (clasificacion_suelo or '').strip().upper()
                _sandy_cls8 = {'SM', 'SP', 'SW', 'GW', 'GP', 'GM', 'GC', 'SC',
                                'SP-SM', 'SW-SM', 'SW-SC', 'SP-SC',
                                'GW-GM', 'GP-GM', 'GW-GC', 'GP-GC'}
                _firm_clay8 = {'CL', 'CH', 'CL-ML', 'CL-CH'}
                if _cls_up8 in _sandy_cls8:
                    _tipo_auto = 3   # Arenoso
                elif _cls_up8 in _firm_clay8:
                    _tipo_auto = 1   # Arcilloso firme
                else:
                    _tipo_auto = 2   # Arcilloso blando (ML, MH, ML-CL, default)

                # 3. N° golpes SPT desde la misma secuencia de correlación geotécnica
                #    Los valores en H10/H11 deben coincidir con los que aparecerán
                #    en la columna F de la plantilla de correlación (mismas profundidades).
                _pisos_cfg_8 = get_pisos_config(pisos_int)
                _spt_seq_8 = get_spt_values(_pisos_cfg_8['template_id'], True)
                # Índice de profundidad ≈ Df: depths = 0.45, 1.45, 2.45, ...
                _df_idx8 = max(0, int(_df_auto - 0.45))
                _n10_auto = int(_spt_seq_8[_df_idx8]) if _df_idx8 < len(_spt_seq_8) else 15
                _n11_auto = int(_spt_seq_8[_df_idx8 + 1]) if (_df_idx8 + 1) < len(_spt_seq_8) else _n10_auto + 5

                # 4. Peso volumétrico Gm típico según tipo de suelo
                _gm_vals8 = [float(lr.get('gamma') or 0) for lr in (perforaciones or []) if float(lr.get('gamma') or 0) > 0]
                if _gm_vals8:
                    _gm_auto = round(sum(_gm_vals8) / len(_gm_vals8), 2)
                elif _tipo_auto == 3:
                    _gm_auto = 1.85   # suelo arenoso
                elif _tipo_auto == 1:
                    _gm_auto = 1.80   # arcilla firme
                else:
                    _gm_auto = 1.74   # arcilla/limo blando

                # 5. c y Φ efectivos drenados según tipo + SPT
                #    Correlación: consistencia del suelo → c' y Φ' (Terzaghi-Peck / Peck et al.)
                _n_rep8 = _n10_auto
                if _tipo_auto == 3:   # Arenoso: c'≈0, Φ' desde SPT (Peck et al.)
                    _c_auto = 0.0
                    if _n_rep8 < 10:
                        _phi_auto = 28.0
                    elif _n_rep8 < 20:
                        _phi_auto = 30.0
                    elif _n_rep8 < 30:
                        _phi_auto = 33.0
                    elif _n_rep8 < 50:
                        _phi_auto = 36.0
                    else:
                        _phi_auto = 40.0
                elif _tipo_auto == 1:  # Arcilloso firme (CL, CH)
                    if _n_rep8 < 4:
                        _c_auto, _phi_auto = 0.08, 12.0
                    elif _n_rep8 < 8:
                        _c_auto, _phi_auto = 0.12, 16.0
                    elif _n_rep8 < 15:
                        _c_auto, _phi_auto = 0.18, 22.0
                    elif _n_rep8 < 30:
                        _c_auto, _phi_auto = 0.25, 26.0
                    else:
                        _c_auto, _phi_auto = 0.35, 30.0
                else:                  # Arcilloso blando (ML, MH, ML-CL, default)
                    if _n_rep8 < 4:
                        _c_auto, _phi_auto = 0.05, 8.0
                    elif _n_rep8 < 8:
                        _c_auto, _phi_auto = 0.08, 12.0
                    elif _n_rep8 < 15:
                        _c_auto, _phi_auto = 0.12, 18.0
                    elif _n_rep8 < 30:
                        _c_auto, _phi_auto = 0.18, 22.0
                    else:
                        _c_auto, _phi_auto = 0.25, 26.0

                # 6. Dimensiones del cimiento (B, L) según pisos
                if pisos_int <= 2:
                    _b_auto = _l_auto = 1.0
                elif pisos_int <= 5:
                    _b_auto = _l_auto = 1.3
                elif pisos_int <= 10:
                    _b_auto = _l_auto = 1.5
                else:
                    _b_auto = _l_auto = 2.0

                # 7. Aplicar (override del usuario tiene prioridad sobre el auto-cálculo)
                _cp = capacidad_portante or {}
                _n10 = int(_cp['n10'])        if _cp.get('n10')         is not None else _n10_auto
                _n11 = int(_cp['n11'])        if _cp.get('n11')         is not None else _n11_auto
                _df  = float(_cp['df'])       if _cp.get('df')          is not None else _df_auto
                _gm  = float(_cp['gm'])       if _cp.get('gm')          is not None else _gm_auto
                _c   = float(_cp['cohesion_c']) if _cp.get('cohesion_c') is not None else _c_auto
                _phi = float(_cp['angulo_fi']) if _cp.get('angulo_fi')  is not None else _phi_auto
                _b   = float(_cp['ancho_b'])  if _cp.get('ancho_b')     is not None else _b_auto
                _l   = float(_cp['largo_l'])  if _cp.get('largo_l')     is not None else _l_auto
                _tipo = int(_cp['tipo_suelo']) if _cp.get('tipo_suelo') is not None else _tipo_auto

                self._set_cell_value(target_sheet, 'H10', _n10)
                self._set_cell_value(target_sheet, 'H11', _n11)
                self._set_cell_value(target_sheet, 'H12', round(_df, 2))
                self._set_cell_value(target_sheet, 'H13', round(_gm, 2))
                self._set_cell_value(target_sheet, 'H14', round(_c, 3))
                self._set_cell_value(target_sheet, 'H15', round(_phi, 1))
                self._set_cell_value(target_sheet, 'H16', round(_b, 2))
                self._set_cell_value(target_sheet, 'H17', round(_l, 2))
                self._set_cell_value(target_sheet, 'H18', _tipo)

                wb_tmp.save(work_file)
                wb_tmp.close()
            finally:
                try:
                    self._remove_calcchain(work_file)
                except Exception:
                    logger.debug("No se pudo eliminar calcChain del libro generado", exc_info=True)

            logger.info("Excel generado exitosamente: %s", work_file)
            return work_file

        laboratorio_template_ids = {'4', '5', '6', '7'}
        if str(template_id) in laboratorio_template_ids:
            try:
                wb_tmp = load_workbook(work_file)
                target_sheet = None
                for sheet_name in wb_tmp.sheetnames:
                    if self._normalize(sheet_name) == self._normalize('Clasificacion'):
                        target_sheet = wb_tmp[sheet_name]
                        break
                if target_sheet is None:
                    target_sheet = wb_tmp[wb_tmp.sheetnames[0]]

                project_value = data.get('proyecto_ubicacion')
                cliente_value = self._format_client(data.get('cliente'))
                fecha_original_value = data.get('fecha_registro_original', data.get('fecha_registro'))
                fecha_plus_20_value = data.get('fecha_registro')
                laboratorio_layer_map = {'4': 0, '5': 1, '6': 2, '7': 3}
                layer_index = laboratorio_layer_map.get(str(template_id))
                selected_layer = (perforaciones or [])[layer_index] if layer_index is not None and len(perforaciones or []) > layer_index else None
                layer_type_text = None
                if selected_layer:
                    layer_type_text = self._clean_soil_text(
                        selected_layer.get('tipo_suelo_principal'),
                        selected_layer.get('color_predominante'),
                    )
                    if not layer_type_text:
                        layer_type_text = self._clean_soil_text(
                            selected_layer.get('descripcion_suelo'),
                            selected_layer.get('color_predominante'),
                        )
                    if layer_type_text:
                        layer_type_text = layer_type_text.upper()
                self._set_cell_value(target_sheet, 'H2', project_value)
                self._set_cell_value(target_sheet, 'L2', fecha_plus_20_value)
                self._set_cell_value(target_sheet, 'Q4', fecha_original_value)
                self._set_cell_value(target_sheet, 'E5', cliente_value)
                if layer_type_text:
                    self._set_cell_value(target_sheet, 'E6', layer_type_text)

                # M4 = max(1, int(profundidad_z)) for the layer assigned to THIS lab.
                # Mapping: Lab4→capa1, Lab5→capa2, Lab6→capa3, Lab7→capa4.
                # 0.45 → int=0 → max(1,0)=1 | 1.45→1 | 2.45→2 | 3.45→3 ...
                # If the lab's capa doesn't exist, fall back to the last available capa.
                _m4_source = selected_layer if selected_layer is not None else (perforaciones[-1] if perforaciones else None)
                _m4_value = None
                if _m4_source is not None:
                    try:
                        _m4_value = max(1, int(float(_m4_source.get('profundidad_z', 0))))
                    except (ValueError, TypeError):
                        pass
                if not _m4_value:
                    _m4_value = 6 if pisos_int <= 3 else (15 if pisos_int <= 10 else 25)
                self._set_cell_value(target_sheet, 'M4', _m4_value)
                self._set_cell_value(target_sheet, 'Q5', _m4_value)

                _atterberg = self._atterberg_for_clasificacion(clasificacion_suelo or '')
                _k14_used = None
                _has_cls = bool((clasificacion_suelo or '').strip())
                _lp_raw: list = []

                # Manual lab values provided by the user via the frontend form
                _vl = valores_laboratorio or {}
                _manual_ll = _vl.get('limite_liquido')   # float | None
                _manual_lp = _vl.get('limite_plastico')  # float | None
                _manual_humedad = _vl.get('humedad')     # float | None

                # --- Casagrande data generation ---
                # Priority: classification > manual values > auto-random.
                # F30/F31 are formula cells (=ROUND(Q27,0) / =ROUND(Q28,0)) that
                # compute LL/LP from the Casagrande inputs automatically.
                # When a classification IS selected we must NOT overwrite F30/F31,
                # because the formula already produces the correct LL/LP for that
                # classification.  Writing a conflicting value (e.g. LL=45.51 for
                # MH which needs LL>50) would make the USCS formula give ML instead.
                # Manual LL/LP are only written to F30/F31 when there is NO
                # classification, so the user's exact decimal is used as-is.
                if _has_cls and _atterberg:
                    # Classification wins — generate Casagrande data that satisfies it.
                    _k14_used, _lp_raw = self._apply_atterberg_to_sheet(target_sheet, _atterberg[0], _atterberg[1])
                    if (clasificacion_suelo or '').strip().upper() in ('SM', 'C'):
                        self._set_sm_gradation(target_sheet, template_id)
                elif _manual_ll is not None or _manual_lp is not None:
                    # No classification — manual values drive Casagrande generation.
                    _ll = float(_manual_ll) if _manual_ll is not None else float(random.randint(35, 45))
                    _lp = float(_manual_lp) if _manual_lp is not None else float(random.randint(15, 25))
                    _k14_used, _lp_raw = self._apply_atterberg_to_sheet(target_sheet, _ll, _lp)
                else:
                    # Default: random Casagrande blow counts (3-point test)
                    _k14_used = random.choice([14, 15, 16])
                    _l14 = random.choice([24, 25, 26])
                    _m14 = random.choice([34, 35, 36])
                    self._set_cell_value(target_sheet, 'K14', _k14_used)
                    self._set_cell_value(target_sheet, 'L14', _l14)
                    self._set_cell_value(target_sheet, 'M14', _m14)
                    _ll_est = random.randint(35, 45)
                    for _n, (_wet_ref, _dry_ref, _tara_ref) in [
                        (_k14_used, ('K15', 'K16', 'K17')),
                        (_l14,      ('L15', 'L16', 'L17')),
                        (_m14,      ('M15', 'M16', 'M17')),
                    ]:
                        _moist = _ll_est * (25 / _n) ** 0.121
                        _tara = self._random_tara_ll()
                        _dry_t = self._random_nonzero_decimal(_tara + 9.0, min(24.9, _tara + 18.0))
                        _wet_t = round(_dry_t + (_dry_t - _tara) * _moist / 100, 2)
                        self._set_cell_value(target_sheet, _tara_ref, _tara)
                        self._set_cell_value(target_sheet, _dry_ref, _dry_t)
                        self._set_cell_value(target_sheet, _wet_ref, _wet_t)
                    _lp_est = random.randint(15, 25)
                    for _wet_r, _dry_r, _tara_r in [
                        ('K22', 'K23', 'K24'),
                        ('L22', 'L23', 'L24'),
                        ('M22', 'M23', 'M24'),
                    ]:
                        _tara = self._random_tara_ll()
                        _dry_t = self._random_nonzero_decimal(_tara + 5.0, min(24.9, _tara + 14.0))
                        _wet_t = round(_dry_t + (_dry_t - _tara) * _lp_est / 100, 2)
                        self._set_cell_value(target_sheet, _tara_r, _tara)
                        self._set_cell_value(target_sheet, _dry_r, _dry_t)
                        self._set_cell_value(target_sheet, _wet_r, _wet_t)
                        if len(_lp_raw) < 2:
                            _lp_raw.append([_wet_t, _dry_t, _tara])

                # --- Summary cell overrides ---
                # F30/F31 only when there is NO classification (classification formula
                # handles them automatically via =ROUND(Q27,0) / =ROUND(Q28,0)).
                # M25 (humidity) is independent of classification, always applied.
                _f30_val = _manual_ll if not _has_cls else None
                _f31_val = _manual_lp if not _has_cls else None
                for _cell_ref, _val in [
                    ('F30', _f30_val),
                    ('F31', _f31_val),
                    ('M25', _manual_humedad),
                ]:
                    if _val is not None:
                        self._set_cell_value(target_sheet, _cell_ref, float(_val))
                        _resolved = _cell_ref
                        if target_sheet[_cell_ref].__class__.__name__ == 'MergedCell':
                            _col_l, _row_n = coordinate_from_string(_cell_ref)
                            _col_n = column_index_from_string(_col_l)
                            for _mr in target_sheet.merged_cells.ranges:
                                if _mr.min_row <= _row_n <= _mr.max_row and _mr.min_col <= _col_n <= _mr.max_col:
                                    _resolved = _mr.start_cell.coordinate
                                    break
                        target_sheet[_resolved].number_format = '0.##'

                # Tara N° del ensayo LP — IDs de los recipientes (K21/L21/M21).
                _tara_id_1 = random.randint(10, 60)
                _tara_id_2 = random.randint(10, 60)
                _tara_id_3 = random.randint(61, 99)
                try:
                    self._set_cell_value(target_sheet, 'K21', _tara_id_1)
                    self._set_cell_value(target_sheet, 'L21', _tara_id_2)
                    self._set_cell_value(target_sheet, 'M21', _tara_id_3)
                except Exception:
                    pass

                # Tara N° del ensayo LL — IDs de los recipientes (K13/L13/M13).
                # Estos mismos valores van a C12/D12 del INCONFINADO.
                _ll_tara_id_1 = random.randint(61, 99)
                _ll_tara_id_2 = random.randint(10, 60)
                _ll_tara_id_3 = random.randint(10, 60)
                try:
                    self._set_cell_value(target_sheet, 'K13', _ll_tara_id_1)
                    self._set_cell_value(target_sheet, 'L13', _ll_tara_id_2)
                    self._set_cell_value(target_sheet, 'M13', _ll_tara_id_3)
                except Exception:
                    pass

                # Save lab state from the first lab template so INCONFINADO and
                # ASENTAMIENTOS can use consistent values (n_golpes, lp, humedad).
                if str(template_id) == '4':
                    # When classification is set, use its LP; otherwise use manual or default.
                    _lab_lp = (
                        float(_atterberg[1]) if _has_cls and _atterberg
                        else (float(_manual_lp) if _manual_lp is not None else 22)
                    )
                    _lab_n = _k14_used if _k14_used is not None else 14
                    _lab_state: dict = {'n_golpes': _lab_n, 'lp': float(_lab_lp)}
                    if _manual_humedad is not None:
                        _lab_state['humedad'] = float(_manual_humedad)
                    # Persist raw LP measurements so INCONFINADO reuses the same weights.
                    if _lp_raw:
                        _lab_state['lp_raw'] = _lp_raw
                    # Persist LL tara IDs (K13/L13) → INCONFINADO C12/D12.
                    _lab_state['tara_ids'] = [_ll_tara_id_1, _ll_tara_id_2]
                    self._save_lab_state(_lab_state)

                wb_tmp.save(work_file)
                wb_tmp.close()
            finally:
                try:
                    self._remove_calcchain(work_file)
                except Exception:
                    logger.debug("No se pudo eliminar calcChain del libro generado", exc_info=True)

            logger.info("Excel generado exitosamente: %s", work_file)
            return work_file

        if str(template_id) == '9':
            try:
                wb_tmp = load_workbook(work_file)
                target_sheet = wb_tmp.active

                cliente_value = self._format_client(data.get('cliente'))
                fecha_value = data.get('fecha_registro')
                proyecto_text = data.get('proyecto_ubicacion')
                proyecto_value, ubicacion_value = self._split_project_location(proyecto_text)
                sondeo_value = str(data.get('sondeo') or 'P-1').strip().upper()
                nivel_freatico_value = str(data.get('nivel_freatico') or 'N.A.').strip()

                # Fila 7: cliente y fecha
                self._set_cell_value(target_sheet, 'C7', cliente_value)
                self._set_cell_value(target_sheet, 'F7', 'FECHA:')
                self._set_cell_value(target_sheet, 'G7', fecha_value)

                # Fila 8: proyecto | sondeo MUESTRA N° | nivel freático
                self._set_cell_value(target_sheet, 'C8', proyecto_value or str(proyecto_text or '').strip().upper())
                self._set_cell_value(target_sheet, 'F8', sondeo_value)
                self._set_cell_value(target_sheet, 'G8', 'MUESTRA')
                self._set_cell_value(target_sheet, 'H8', 1)
                self._set_cell_value(target_sheet, 'J8', nivel_freatico_value)

                # Fila 9: localización — municipio del informe Word (si lo seleccionó el usuario)
                _municipio9 = data.get('municipio_word') or ubicacion_value
                self._set_cell_value(target_sheet, 'C9', _municipio9)

                # Fila 10: descripción de la muestra (se rellena abajo con perforaciones)
                self._set_cell_value(target_sheet, 'C10', None)

                selected_description = None
                if perforaciones:
                    first_layer = perforaciones[0]
                    selected_description = self._clean_soil_text(
                        first_layer.get('descripcion_suelo'),
                        first_layer.get('color_predominante'),
                    )
                    if not selected_description:
                        selected_description = self._clean_soil_text(
                            first_layer.get('tipo_suelo_principal'),
                            first_layer.get('color_predominante'),
                        )
                    if selected_description:
                        selected_description = selected_description.upper()

                if selected_description:
                    self._set_cell_value(target_sheet, 'C10', selected_description)

                # Mirror moisture content data from LABORATORIO LP test.
                # C13/D13 = húmedo+tara, C14/D14 = seco+tara, C15/D15 = tara.
                # C16/D16 and C17 are formula cells — leave them untouched.
                # When lp_raw is present (saved by template 4), reuse those exact weights
                # so laboratorio and inconfinado show consistent moisture data.
                _lab_st9 = self._load_lab_state()
                _lp_val = int(_lab_st9.get('lp', 22)) if _lab_st9 else 22
                _lp_raw9 = (_lab_st9.get('lp_raw') or []) if _lab_st9 else []
                for _idx, (_ch, _cs, _ct) in enumerate([('C13', 'C14', 'C15'), ('D13', 'D14', 'D15')]):
                    if _idx < len(_lp_raw9):
                        _humid_t, _seco_t, _tara = _lp_raw9[_idx]
                    else:
                        _tara = round(random.uniform(5.5, 8.0), 2)
                        _seco_net = round(random.uniform(11.0, 16.0), 2)
                        _seco_t = round(_tara + _seco_net, 2)
                        _humid_t = round(_seco_t + _seco_net * _lp_val / 100, 2)
                    self._set_cell_value(target_sheet, _ch, _humid_t)
                    self._set_cell_value(target_sheet, _cs, _seco_t)
                    self._set_cell_value(target_sheet, _ct, _tara)
                # C12/D12 = Tara N° — mismos IDs que K21/L21 del laboratorio (template 4)
                _tara_ids9 = (_lab_st9.get('tara_ids') or []) if _lab_st9 else []
                _c12 = _tara_ids9[0] if len(_tara_ids9) > 0 else random.randint(10, 60)
                _d12 = _tara_ids9[1] if len(_tara_ids9) > 1 else random.randint(10, 60)
                self._set_cell_value(target_sheet, 'C12', _c12)
                self._set_cell_value(target_sheet, 'D12', _d12)

                # Quitar la imagen "MUESTRA FALLADA" embebida en la plantilla
                target_sheet._images = []

                wb_tmp.save(work_file)
                wb_tmp.close()
            finally:
                try:
                    self._remove_calcchain(work_file)
                except Exception:
                    logger.debug("No se pudo eliminar calcChain del libro generado", exc_info=True)

            logger.info("Excel generado exitosamente: %s", work_file)
            return work_file

        asentamientos_template_ids = {'10', '11'}
        if str(template_id) in asentamientos_template_ids:
            try:
                wb_tmp = load_workbook(work_file)
                target_sheet = wb_tmp.active

                cliente_value = self._format_client(data.get('cliente'))
                fecha_value = data.get('fecha_registro')
                proyecto_text = data.get('proyecto_ubicacion')
                proyecto_value = str(proyecto_text).strip().upper() if proyecto_text is not None else None
                pisos_value = data.get('pisos')
                try:
                    pisos_int = int(pisos_value or 0)
                except Exception:
                    pisos_int = 0

                if pisos_int > 0:
                    pisos_text = str(pisos_int)
                else:
                    pisos_text = ''

                if pisos_int > 0:
                    if pisos_int == 1:
                        e4_updated = 'ESTRUCTURA DE 1 NIVEL'
                    else:
                        e4_updated = f'ESTRUCTURA DE {pisos_int} NIVELES'
                else:
                    e4_updated = target_sheet['E4'].value

                self._set_cell_value(target_sheet, 'E4', e4_updated)
                self._set_cell_value(target_sheet, 'E5', proyecto_value)
                self._set_cell_value(target_sheet, 'E6', cliente_value)
                self._set_cell_value(target_sheet, 'E8', fecha_value)

                # ── Auto-derivar parámetros de asentamiento (misma lógica que template 8) ──

                # 1. Tipo de suelo desde clasificacion_suelo
                _cls_up10 = (clasificacion_suelo or '').strip().upper()
                _sandy_cls10 = {'SM', 'SP', 'SW', 'GW', 'GP', 'GM', 'GC', 'SC',
                                 'SP-SM', 'SW-SM', 'SW-SC', 'SP-SC',
                                 'GW-GM', 'GP-GM', 'GW-GC', 'GP-GC'}
                _firm_clay10 = {'CL', 'CH', 'CL-ML', 'CL-CH'}
                if _cls_up10 in _sandy_cls10:
                    _tipo10 = 3
                elif _cls_up10 in _firm_clay10:
                    _tipo10 = 1
                else:
                    _tipo10 = 2  # arcilloso blando / default

                # 2. Profundidad Df y dimensiones B/L según pisos (idéntico a template 8)
                if pisos_int <= 1:
                    _df10 = 1.0
                elif pisos_int <= 3:
                    _df10 = 1.5
                elif pisos_int <= 6:
                    _df10 = 2.0
                else:
                    _df10 = 2.5

                if pisos_int <= 2:
                    _b10 = _l10 = 1.0
                elif pisos_int <= 5:
                    _b10 = _l10 = 1.3
                elif pisos_int <= 10:
                    _b10 = _l10 = 1.5
                else:
                    _b10 = _l10 = 2.0

                # 3. N° golpes SPT en la misma profundidad de Df (correlación geotécnica)
                _pisos_cfg_10 = get_pisos_config(pisos_int)
                _spt_seq_10 = get_spt_values(_pisos_cfg_10['template_id'], True)
                _df_idx10 = max(0, int(_df10 - 0.45))
                _n19_auto = int(_spt_seq_10[_df_idx10]) if _df_idx10 < len(_spt_seq_10) else 15

                # 4. Carga aplicada qu según plantilla (300 kN ó 800 kN)
                _qu = 300.0 if str(template_id) == '10' else 800.0

                # 5. Módulo de elasticidad E según tipo de suelo + SPT
                #    Correlaciones: Schultze-Melzer (arenas) / Terzaghi-Peck (arcillas)
                _n_e = _n19_auto
                if _tipo10 == 3:   # Arenoso
                    _e_auto = min(60000, max(20000, 1500 * _n_e))
                elif _tipo10 == 1:  # Arcilloso firme
                    _e_auto = min(30000, max(10000, 800 * _n_e))
                else:               # Arcilloso blando
                    _e_auto = min(15000, max(5000, 600 * _n_e))

                # 6. Override del usuario si existe capacidad_portante
                _cp_as = capacidad_portante or {}
                _n19 = int(_cp_as['n10']) if _cp_as.get('n10') is not None else _n19_auto
                _b19 = float(_cp_as['ancho_b']) if _cp_as.get('ancho_b') is not None else _b10
                _l19 = float(_cp_as['largo_l']) if _cp_as.get('largo_l') is not None else _l10

                self._set_cell_value(target_sheet, 'H19', _n19)
                self._set_cell_value(target_sheet, 'H20', _qu)
                self._set_cell_value(target_sheet, 'H21', _e_auto)
                self._set_cell_value(target_sheet, 'H24', round(_b19, 2))
                self._set_cell_value(target_sheet, 'H25', round(_l19, 2))

                wb_tmp.save(work_file)
                wb_tmp.close()
            finally:
                try:
                    self._remove_calcchain(work_file)
                except Exception:
                    logger.debug("No se pudo eliminar calcChain del libro generado", exc_info=True)

            logger.info("Excel generado exitosamente: %s", work_file)
            return work_file

        # Include template '15' as legacy (.xls) because it maps to P-3.xls
        # which is an old Excel BIFF file; attempting to treat it as a ZIP
        # will raise "File is not a zip file". Keep legacy handling here.
        legacy_xls_template_ids = {'12', '13', '14', '15'}
        if str(template_id) in legacy_xls_template_ids:
            try:
                fecha_value = data.get('fecha_registro', data.get('fecha_registro_original'))
                project_value = f"PROYECTO: {data.get('proyecto_ubicacion', '')}".strip()
                _pisos_cfg = get_pisos_config(pisos_int)
                e5_value = _pisos_cfg["max_depth"]
                n_campo_values = self._get_legacy_n_campo_values(template_id, use_lower, pisos_int)

                # Build soil descriptions, colors, and depth values for column I of P-1.xls.
                # Mirrors the proportional layer order used for column Q of plantilla_1.xlsx.
                _soil_descs: List[str] = []
                _soil_colors: List[str] = []
                _depth_vals: List[float] = []
                for _layer in (perforaciones or []):
                    _desc = self._clean_soil_text(
                        _layer.get('descripcion_suelo'), _layer.get('color_predominante')
                    )
                    if not _desc:
                        _desc = self._clean_soil_text(
                            _layer.get('tipo_suelo_principal'), _layer.get('color_predominante')
                        )
                    _soil_descs.append(_desc or "")
                    _soil_colors.append(_layer.get('color_predominante') or "")
                    try:
                        _d = float(_layer.get('profundidad_z') or 0)
                    except Exception:
                        _d = 0.0
                    _depth_vals.append(_d)

                # P-1 → Lab 1 ('4'), P-2 → Lab 2 ('5'), P-3 → Lab 3 ('6'), P-4 → Lab 4 ('7')
                # ALL lab data (w, LL, LP, IP, gamma, USCS) goes ONLY to the last layer row
                # (the representative sample), matching the original template structure.
                _p_to_lab = {'12': '4', '13': '5', '14': '6', '15': '7'}
                _this_lab_id = _p_to_lab.get(str(template_id))

                _n_layers = max(len(perforaciones or []), 7)
                _lab_data_per_layer: List[Optional[dict]] = [None] * _n_layers

                if _this_lab_id:
                    _vl = (valores_laboratorio_por_lab or {}).get(_this_lab_id) or valores_laboratorio
                    _cls = (clasificaciones_por_lab or {}).get(_this_lab_id) or clasificacion_suelo
                    _ll = _lp = _hum = None
                    try:
                        _ll = float(_vl['limite_liquido']) if _vl and _vl.get('limite_liquido') is not None else None
                    except (ValueError, TypeError):
                        pass
                    try:
                        _lp = float(_vl['limite_plastico']) if _vl and _vl.get('limite_plastico') is not None else None
                    except (ValueError, TypeError):
                        pass
                    try:
                        _hum = float(_vl['humedad']) if _vl and _vl.get('humedad') is not None else None
                    except (ValueError, TypeError):
                        pass
                    _ip = round(_ll - _lp, 2) if _ll is not None and _lp is not None else None
                    _last_idx = max(0, len(perforaciones or []) - 1)
                    _gam_last = None
                    if perforaciones and _last_idx < len(perforaciones):
                        try:
                            _gam_last = float(perforaciones[_last_idx].get('gamma') or 0) or None
                        except (ValueError, TypeError):
                            pass
                    if _last_idx < _n_layers:
                        _lab_data_per_layer[_last_idx] = {
                            'humedad': _hum,
                            'limite_liquido': _ll,
                            'limite_plastico': _lp,
                            'indice_plasticidad': _ip,
                            'gamma': _gam_last,
                            'clasificacion_uscs': _cls,
                        }

                # All legacy .xls templates (P-1 through P-4) use the same dynamic
                # depth scale expansion and proportional column I color logic.
                _expand_levels = _pisos_cfg["expand_levels"]
                _municipio_val = str(data.get('municipio_word') or '').strip().upper() or None
                self._fill_legacy_xls_template(
                    work_file, project_value, fecha_value, e5_value, n_campo_values,
                    soil_descriptions=_soil_descs if _soil_descs else None,
                    soil_colors=_soil_colors if _soil_colors else None,
                    depth_values=_depth_vals if _depth_vals else None,
                    expand_depth_levels=_expand_levels,
                    photo_paths=photo_paths,
                    lab_data_per_layer=_lab_data_per_layer if any(x is not None for x in _lab_data_per_layer) else None,
                    municipio=_municipio_val,
                )
            except Exception:
                try:
                    if work_file.exists():
                        work_file.unlink()
                except Exception:
                    logger.debug("No se pudo limpiar el archivo .xls temporal", exc_info=True)
                raise

            logger.info("Excel generado exitosamente: %s", work_file)
            return work_file

        sheet_targets = self._resolve_sheet_targets(work_file)
        sheet_updates = {}

        primary_updates = {}
        field_mapping = get_field_mapping(template_id)
        for field_name, cell_refs in field_mapping.items():
            value = data.get(field_name)
            if value is None:
                continue

            if isinstance(value, date):
                value = value.strftime("%d/%m/%Y")

            for cell_ref in cell_refs:
                primary_updates[cell_ref] = value

        perforacion_mapping = get_perforacion_mapping(template_id)
        start_row = perforacion_mapping["tabla_inicio_row"]
        allowed_columns = perforacion_mapping["allowed_columns"]
        if str(template_id) == "1":
            f_end = 16
        elif str(template_id) == "2":
            f_end = 25
        elif str(template_id) == "3":
            f_end = 35
        else:
            f_end = start_row + max(0, len(perforaciones or [])) - 1

        rng = random.SystemRandom()
        reduce_next = bool(rng.getrandbits(1))

        for index, row_data in enumerate(perforaciones or []):
            row_number = start_row + index
            for field_name, column_letter in allowed_columns.items():
                if field_name == "profundidad_z":
                    continue
                if field_name == "descripcion_suelo":
                    continue
                if field_name == "gamma":
                    # gamma is written per layer segment in the segment loop below,
                    # not one value per perforacion row.
                    continue

                value = row_data.get(field_name, "")
                if field_name == "n_campo_spt" and 10 <= row_number <= 16 and isinstance(value, int):
                    value = max(0, value - 1) if reduce_next else value
                    reduce_next = not reduce_next

                primary_updates[f"{column_letter}{row_number}"] = value

        # If the soil description column is empty, mirror the same-row L cell.
        for row_number in range(start_row, f_end + 1):
            p_cell = f"P{row_number}"
            if primary_updates.get(p_cell) in (None, ""):
                primary_updates[p_cell] = f"=L{row_number}"

        # Enforce specific SPT values per plantilla
        spt_final_values = self._calculate_spt_values(template_id, use_lower, pisos_int)
        
        # Apply final SPT values to Column F and persist so paired .xls templates
        # (P-1/P-2/P-3/P-4) can mirror the exact same values in their T column.
        if str(template_id) in ("1", "2", "3"):
            spt_start = 10
            for i, v in enumerate(spt_final_values):
                primary_updates[f"F{spt_start + i}"] = v
            self._save_spt_state(str(template_id), spt_final_values)
            # Column A (Z(m)): auto-generate depth sequence 0.45, 1.45, 2.45, …
            for _ai in range(f_end - start_row + 1):
                primary_updates[f"A{start_row + _ai}"] = round(0.45 + _ai * 1.0, 2)
        if primary_updates:
            primary_sheet = sheet_targets.get(self._normalize("P3"))
            if primary_sheet:
                sheet_updates[primary_sheet] = primary_updates

        parametros_updates = {}
        parametros_mapping = get_parametros_mapping(template_id)
        for index, row_data in enumerate(parametros or []):
            row_number = parametros_mapping["tabla_inicio_row"] + index
            for field_name, column_letter in parametros_mapping["allowed_columns"].items():
                parametros_updates[f"{column_letter}{row_number}"] = row_data.get(field_name, "")

        if parametros_updates:
            parametros_sheet = sheet_targets.get(self._normalize("Parametros"))
            if parametros_sheet:
                sheet_updates[parametros_sheet] = parametros_updates

        if sheet_updates:
            # Simpler approach: load the copied template with openpyxl, apply updates and save.
            try:
                wb_tmp = load_workbook(work_file)
                # reverse map from target path to normalized sheet name
                target_to_sheetname = {v.lstrip('/'): k for k, v in sheet_targets.items()}
                for target_path, updates in sheet_updates.items():
                    normalized_target = target_path.lstrip('/')
                    norm_sheet = target_to_sheetname.get(normalized_target)
                    if not norm_sheet:
                        # try matching without xl/ prefix
                        norm_sheet = target_to_sheetname.get(normalized_target.replace('xl/',''))
                    if not norm_sheet:
                        continue
                    # find actual sheet by normalized name
                    sheet_obj = None
                    for name in wb_tmp.sheetnames:
                        if self._normalize(name) == norm_sheet:
                            sheet_obj = wb_tmp[name]
                            break
                    if not sheet_obj:
                        continue
                    for cell_ref, value in updates.items():
                        try:
                            # write into merged anchors if necessary
                            if ':' in cell_ref:
                                cell_ref = cell_ref.split(':')[0]
                            cell = sheet_obj[cell_ref]
                            if cell.__class__.__name__ == 'MergedCell':
                                for merged_range in sheet_obj.merged_cells.ranges:
                                    if cell_ref in merged_range:
                                        sheet_obj[merged_range.start_cell.coordinate] = value
                                        break
                            else:
                                sheet_obj[cell_ref] = value
                        except Exception:
                            continue

                    def build_soil_text(layer: dict) -> str:
                        descripcion = self._clean_soil_text(layer.get('descripcion_suelo'), layer.get('color_predominante'))
                        if descripcion:
                            return descripcion
                        tipo = self._clean_soil_text(layer.get('tipo_suelo_principal'), layer.get('color_predominante'))
                        return tipo

                    if str(template_id) in ("1", "2", "3"):
                        q_segments = self._build_depth_soil_segments_exact(
                            list(perforaciones or []), start_row, f_end
                        )
                    else:
                        q_segments = self._build_soil_segments(sheet_obj, list(perforaciones or []), start_row, f_end)

                    # Coalesce adjacent segments that render the same soil text so they
                    # appear as a single merged block (user requested centrar y combinarlas)
                    def _segment_text(seg):
                        return build_soil_text(seg[2])

                    coalesced = []
                    for seg in q_segments:
                        if not coalesced:
                            coalesced.append(list(seg))
                            continue
                        prev = coalesced[-1]
                        if _segment_text(prev) == _segment_text(seg) and seg[0] == prev[1] + 1:
                            # extend previous segment
                            prev[1] = seg[1]
                        else:
                            coalesced.append(list(seg))

                    # normalize back to tuples
                    q_segments = [(s[0], s[1], s[2]) for s in coalesced]

                    # Unmerge Q and B columns in the data range before reapplying per-segment merges.
                    # Column I must never be touched (Marcuson K factor formulas).
                    # Note: startswith check removed for B — catches multi-col merges like A10:B12
                    # where the anchor is column A, not B.
                    for merged_range in list(sheet_obj.merged_cells.ranges):
                        if merged_range.max_row < start_row or merged_range.min_row > f_end:
                            continue
                        if merged_range.min_col <= 17 <= merged_range.max_col:
                            try:
                                sheet_obj.unmerge_cells(str(merged_range))
                            except Exception:
                                pass
                        elif merged_range.min_col <= 2 <= merged_range.max_col:
                            try:
                                sheet_obj.unmerge_cells(str(merged_range))
                            except Exception:
                                pass

                    # Clear column B values in the full data range AFTER unmerge so all cells
                    # (including formerly-merged ones) are cleared before writing new values.
                    if str(template_id) in ("1", "2", "3"):
                        for _br in range(start_row, f_end + 1):
                            try:
                                sheet_obj.cell(row=_br, column=2).value = None
                            except Exception:
                                pass

                    # Clear Q column values and (for correlation templates) fills so
                    # rows beyond the last layer are completely empty.
                    _no_fill = PatternFill(fill_type=None)
                    for _qr in range(start_row, f_end + 1):
                        try:
                            _qcell = sheet_obj.cell(row=_qr, column=17)
                            _qcell.value = None
                            if str(template_id) in ("1", "2", "3"):
                                _qcell.fill = _no_fill
                        except Exception:
                            pass

                    for seg_idx, (row_start, row_end, layer) in enumerate(q_segments):
                        self._apply_q_segment_format(
                            sheet_obj, row_start, row_end, layer, build_soil_text(layer)
                        )
                        # NOTE: Column I contains critical Marcuson K factor formulas =IF(H#<1,(1.41),(0.92))
                        # NEVER modify Column I - it must remain unchanged with original formulas

                    # ── Column B: gamma values derived from Q's actual merged structure ──
                    # After the Q loop creates the real merges, each merged block in Q
                    # represents one geotechnical layer.  Layer 1 → 15, layer 2 → 16, …
                    # The gamma is repeated in every individual B cell of that row span.
                    if str(template_id) in ("1", "2", "3"):
                        _b_blocks: list = []
                        _b_covered: set = set()
                        for _mr in list(sheet_obj.merged_cells.ranges):
                            if _mr.min_col <= 17 <= _mr.max_col:
                                _rs = max(_mr.min_row, start_row)
                                _re = min(_mr.max_row, f_end)
                                if _rs <= _re:
                                    _b_blocks.append((_rs, _re))
                                    _b_covered.update(range(_rs, _re + 1))
                        # Q rows not part of any merge → own single-row layer
                        for _r in range(start_row, f_end + 1):
                            if _r not in _b_covered:
                                _b_blocks.append((_r, _r))
                        _b_blocks.sort()
                        _white_fill = PatternFill(fill_type="solid", fgColor="FFFFFF")
                        _thin = Side(style="thin")
                        _b_border = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)
                        for _idx, (_rs, _re) in enumerate(_b_blocks):
                            _gamma = 15 + _idx
                            for _row in range(_rs, _re + 1):
                                try:
                                    _cell = sheet_obj.cell(row=_row, column=2)
                                    _cell.value = _gamma
                                    _cell.fill = _white_fill
                                    _cell.border = _b_border
                                    _cell.alignment = Alignment(
                                        horizontal='center', vertical='center'
                                    )
                                except Exception:
                                    pass

                # finally save workbook once after all data updates
                wb_tmp.save(work_file)
                wb_tmp.close()
            except Exception:
                # fallback: try the previous xml-replacement method
                try:
                    with NamedTemporaryFile(delete=False, suffix='.xlsx') as tf:
                        temp_saved = Path(tf.name)
                    wb_tmp = load_workbook(work_file)
                    wb_tmp.save(temp_saved)
                    wb_tmp.close()
                    with ZipFile(temp_saved, 'r') as tmpz:
                        tmp_names = tmpz.namelist()
                        replacements = {}
                        for target_path in sheet_updates.keys():
                            normalized = target_path.lstrip('/')
                            if normalized in tmp_names:
                                replacements[normalized] = tmpz.read(normalized)
                    if replacements:
                        self._write_workbook_preserving_parts(work_file, replacements)
                finally:
                    try:
                        if temp_saved.exists():
                            temp_saved.unlink()
                    except Exception:
                        pass

        # Aplicar todos los recursos visuales registrados para este template.
        # El resource_manager sabe automáticamente qué recurso va en qué documento.
        try:
            resource_manager.apply_to_excel(work_file, str(template_id))
        except Exception:
            logger.debug(
                "No se pudieron aplicar recursos visuales en %s", work_file, exc_info=True
            )

        # remove calcChain if present to avoid Excel repair dialogs
        try:
            self._remove_calcchain(work_file)
        except Exception:
            logger.debug("No se pudo eliminar calcChain del libro generado", exc_info=True)

        logger.info("Excel generado exitosamente: %s", work_file)
        return work_file

    def _remove_calcchain(self, workbook_path: Path) -> None:
        # safe remove of xl/calcChain.xml, its override and workbook rel
        with ZipFile(workbook_path, 'r') as src:
            # Fast-path: openpyxl already strips calcChain on save, so this is
            # usually a no-op. Only read all entries when calcChain is actually present.
            if 'xl/calcChain.xml' not in src.namelist():
                return
            entries = {info.filename: src.read(info.filename) for info in src.infolist()}

        changed = False
        if 'xl/calcChain.xml' in entries:
            entries.pop('xl/calcChain.xml', None)
            changed = True

        # remove override from [Content_Types].xml
        if '[Content_Types].xml' in entries:
            try:
                ct = ET.fromstring(entries['[Content_Types].xml'])
                ns = {'ct': 'http://schemas.openxmlformats.org/package/2006/content-types'}
                for o in list(ct.findall('ct:Override', ns)):
                    if o.attrib.get('PartName') == '/xl/calcChain.xml':
                        ct.remove(o)
                        changed = True
                entries['[Content_Types].xml'] = ET.tostring(ct, encoding='utf-8', xml_declaration=True)
            except Exception:
                pass

        # remove relationship from xl/_rels/workbook.xml.rels
        if 'xl/_rels/workbook.xml.rels' in entries:
            try:
                wr = ET.fromstring(entries['xl/_rels/workbook.xml.rels'])
                for rel in list(wr):
                    if 'calcChain' in rel.attrib.get('Type', ''):
                        wr.remove(rel)
                        changed = True
                entries['xl/_rels/workbook.xml.rels'] = ET.tostring(wr, encoding='utf-8', xml_declaration=True)
            except Exception:
                pass

        if not changed:
            return

        # write new zip
        with NamedTemporaryFile(delete=False, suffix='.xlsx') as tf:
            temp_path = Path(tf.name)
        try:
            with ZipFile(temp_path, 'w', compression=ZIP_DEFLATED) as zf:
                for name, data in entries.items():
                    zf.writestr(name, data)
            if workbook_path.exists():
                workbook_path.unlink()
            shutil.move(str(temp_path), str(workbook_path))
        finally:
            if temp_path.exists():
                try:
                    temp_path.unlink()
                except Exception:
                    pass

    def verify_templates(self) -> dict:
        templates_status = {}
        for template_id in ["1", "2"]:
            try:
                path = self._get_template_path(template_id)
                templates_status[f"Template{template_id}"] = {
                    "exists": True,
                    "path": str(path),
                }
            except FileNotFoundError:
                templates_status[f"Template{template_id}"] = {
                    "exists": False,
                    "path": str(self.templates_dir / settings.TEMPLATES_CONFIG.get(template_id, f"plantilla_{template_id}.xlsx")),
                }
        return templates_status


# Instancia global del servicio
excel_service = ExcelService()
