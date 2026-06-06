#!/usr/bin/env python3
"""
Script to clean Column I from plantilla_1.xlsx, plantilla_2.xlsx, plantilla_3.xlsx
"""
from pathlib import Path
from openpyxl import load_workbook

template_dir = Path(r"C:\Users\Paola\OneDrive - Tecnologico de Antioquia Institucion Universitaria\Escritorio\Automatización geotecnica\automatizacion_geotecnica\backend\templates\excel")

templates = [
    template_dir / "plantilla_1.xlsx",
    template_dir / "plantilla_2.xlsx",
    template_dir / "plantilla_3.xlsx",
]

print("=" * 80)
print("CLEANING COLUMN I FROM TEMPLATES")
print("=" * 80)

for template_path in templates:
    if not template_path.exists():
        print(f"\n[ERROR] Template not found: {template_path}")
        continue
    
    print(f"\n[INFO] Opening: {template_path.name}")
    
    try:
        # Load the workbook
        workbook = load_workbook(template_path)
        worksheet = workbook.active
        
        # Find all data in Column I
        rows_with_data = []
        for row in worksheet.iter_rows(min_col=9, max_col=9, values_only=False):
            for cell in row:
                if cell.value is not None:
                    rows_with_data.append(cell.row)
        
        print(f"  Found data in Column I at rows: {rows_with_data}")
        
        # Clean all cells in Column I
        rows_cleaned = 0
        for row in worksheet.iter_rows(min_col=9, max_col=9):
            for cell in row:
                if cell.value is not None:
                    cell.value = None
                    rows_cleaned += 1
        
        print(f"  Cleaned {rows_cleaned} cells in Column I")
        
        # Save the workbook
        workbook.save(template_path)
        print(f"  [OK] Saved: {template_path.name}")
        workbook.close()
        
    except Exception as e:
        print(f"  [ERROR] Error processing {template_path.name}: {e}")
        import traceback
        traceback.print_exc()

print("\n" + "=" * 80)
print("CLEANING COMPLETE")
print("=" * 80)
