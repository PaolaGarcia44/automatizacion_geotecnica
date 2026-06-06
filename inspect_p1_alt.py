#!/usr/bin/env python3
"""
Script to inspect Column I using alternative methods
"""
import sys
import zipfile
from pathlib import Path
import xml.etree.ElementTree as ET

p1_file = Path(r"C:\Users\Paola\OneDrive - Tecnologico de Antioquia Institucion Universitaria\Escritorio\Automatización geotecnica\automatizacion_geotecnica\backend\generated\TEST_FRONTEND_EXTRAIDO\P-1.xls")

print(f"[INFO] File exists: {p1_file.exists()}")
print(f"[INFO] File size: {p1_file.stat().st_size if p1_file.exists() else 'N/A'} bytes")
print(f"[INFO] File format: .xls (legacy binary format - requires special handling)")

# Try with openpyxl for .xlsx files
try:
    from openpyxl import load_workbook
    wb = load_workbook(str(p1_file))
    ws = wb.active
    print("\n[INFO] Successfully opened with openpyxl")
    print("\nColumn I values:")
    for row in [8, 10, 12, 14]:
        cell = ws[f"I{row}"]
        print(f"  I{row}: {cell.value}")
except Exception as e:
    print(f"\n[ERROR] Could not open with openpyxl: {type(e).__name__}: {e}")
    print("\nNote: P-1.xls is a legacy BIFF file format (.xls) which requires win32com")
    print("      to read properly. openpyxl only supports modern .xlsx format.")
    print("\n[RECOMMENDATION] The file was edited using Excel COM interface on the backend.")
    print("                 To verify the data, please open P-1.xls manually in Excel.")
